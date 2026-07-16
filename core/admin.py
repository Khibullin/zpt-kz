from django.db.models import Prefetch
from django import forms
from django.contrib import admin
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import path, reverse
from django.utils import timezone
from django.utils.html import escape
from django.utils.safestring import mark_safe
from .models import WhatsAppMessageLog

from openpyxl import load_workbook

from catalog.instagram_service import (
    approve_instagram_publication,
    cancel_instagram_publication,
    mark_stuck_instagram_publication_failed,
    queue_instagram_publication_for_processing,
)
from .models import (
    Country,
    Brand,
    CarModel,
    PartCategory,
    BroadcastSettings,
    Request,
    Seller,
    SellerLead,
    SellerLeadContactCandidate,
    SellerLeadPipelineRun,
    Match,
    RequestDispatch,
    Feedback,
    InstagramPublication,
    BuyerContact,
    BuyerVehicle,
    BuyerCategoryInterest,
    BuyerCityInterest,
    ContactConsent,
    BuyerAudience,
    CONTACT_CONSENT_CHANNEL_WHATSAPP,
    CONTACT_CONSENT_PURPOSE_MARKETING,
)
from core.buyer_audience_admin_forms import (
    BuyerAudienceAdminForm,
    format_criteria_details,
    format_criteria_summary,
)
from core.buyer_contact_admin_filters import (
    BuyerActivityFilter,
    BuyerBrandFilter,
    BuyerCategoryFilter,
    BuyerMarketingConsentFilter,
    BuyerModelFilter,
    BuyerPrimaryCityFilter,
    BuyerRequestCountFilter,
    BuyerTransportTypeFilter,
    build_category_summary,
    build_vehicle_summary,
    marketing_consent_label,
)
from core.services.buyer_contact_utils import mask_phone
from core.services.buyer_audience_service import (
    audience_criteria_has_filters,
    preview_buyer_audience,
)


class SellerImportForm(forms.Form):
    file = forms.FileField(label='Excel файл .xlsx')


def normalize_phone(value):
    return ''.join(ch for ch in str(value or '') if ch.isdigit())


def split_values(value):
    if not value:
        return []

    raw = str(value).replace(',', ';')
    return [item.strip() for item in raw.split(';') if item and item.strip()]


def get_cell(row, headers, name, default=''):
    index = headers.get(name)

    if index is None:
        return default

    value = row[index]

    if value is None:
        return default

    return str(value).strip()


def parse_sellers_xlsx(file_obj):
    workbook = load_workbook(file_obj, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return [], ['Файл пустой']

    header_row = rows[0]

    headers = {
        str(value).strip().lower(): index
        for index, value in enumerate(header_row)
        if value
    }

    required = [
        'seller_name',
        'transport_type',
    ]

    errors = []

    for col in required:
        if col not in headers:
            errors.append(f'Нет обязательной колонки: {col}')

    if errors:
        return [], errors

    parsed_rows = []

    for row_number, row in enumerate(rows[1:], start=2):
        if not row or not any(row):
            continue

        seller_name = get_cell(row, headers, 'seller_name')
        whatsapp = normalize_phone(get_cell(row, headers, 'whatsapp'))
        phone2 = normalize_phone(get_cell(row, headers, 'phone2'))
        city = get_cell(row, headers, 'city', 'Алматы')
        market_location = get_cell(row, headers, 'market_location')
        transport_type = get_cell(row, headers, 'transport_type', 'car').lower()
        categories = get_cell(row, headers, 'categories')
        countries = get_cell(row, headers, 'countries')
        brands = get_cell(row, headers, 'brands')
        seller_type = get_cell(row, headers, 'seller_type', 'seller').lower()
        dispatch_priority_raw = get_cell(row, headers, 'dispatch_priority', '1000')
        notes = get_cell(row, headers, 'notes')

        if not seller_name:
            seller_name = f"Seller {row_number}"
            notes = (notes + " | " if notes else "") + "Требует проверки: нет названия"

        if not whatsapp:
            whatsapp = f"NO-WA-{row_number}"
            notes = (notes + " | " if notes else "") + "Требует проверки: нет WhatsApp"

        if not categories:
            categories = 'general_parts'
            notes = (notes + " | " if notes else "") + "Требует проверки: нет категории"

        if not countries:
            countries = 'Multi'
            notes = (notes + " | " if notes else "") + "Требует проверки: нет страны"

        if not brands:
            brands = 'Multi'
            notes = (notes + " | " if notes else "") + "Требует проверки: нет марки"

        if transport_type not in ['car', 'truck']:
            transport_type = 'car'
            notes = (notes + " | " if notes else "") + "Требует проверки: transport_type исправлен на car"

        if seller_type not in ['seller', 'service', 'both']:
            seller_type = 'seller'
            notes = (notes + " | " if notes else "") + "Требует проверки: seller_type исправлен на seller"

        try:
            dispatch_priority = int(dispatch_priority_raw)
        except ValueError:
            dispatch_priority = 1000
            notes = (notes + " | " if notes else "") + "Требует проверки: priority исправлен на 1000"

        parsed_rows.append({
            'row_number': row_number,
            'seller_name': seller_name[:255],
            'whatsapp': whatsapp,
            'phone2': phone2,
            'city': city,
            'market_location': market_location[:255],
            'transport_type': transport_type,
            'categories': categories,
            'countries': countries,
            'brands': brands,
            'seller_type': seller_type,
            'receive_requests': False,
            'is_test_seller': False,
            'dispatch_priority': dispatch_priority,
            'notes': notes,
        })

    return parsed_rows, errors


def find_seller_by_whatsapp(whatsapp):
    target = normalize_phone(whatsapp)

    if not target:
        return None

    for seller in Seller.objects.all():
        if normalize_phone(seller.whatsapp) == target:
            return seller

    return None


def import_seller_row(row):
    seller = find_seller_by_whatsapp(row['whatsapp'])
    created = False

    if seller is None:
        seller = Seller()
        created = True

    category_names = split_values(row.get('categories'))
    country_names = split_values(row.get('countries'))
    brand_names = split_values(row.get('brands'))

    if not category_names:
        category_names = ['general_parts']

    if not country_names:
        country_names = ['Multi']

    if not brand_names:
        brand_names = ['Multi']

    seller.name = row['seller_name'][:255]
    seller.whatsapp = row['whatsapp'][:20]
    seller.phone2 = row['phone2'][:20]
    seller.city = row['city'][:100]
    seller.market_location = row['market_location'][:255]
    seller.transport_type = row['transport_type']
    seller.seller_type = row['seller_type']
    seller.dispatch_priority = row['dispatch_priority']
    seller.notes = row['notes']

    # Безопасность импорта: никто не получает заявки автоматически.
    seller.receive_requests = False
    seller.is_test_seller = False
    seller.is_active = True
    seller.is_paused = False

    seller.category = category_names[0] if category_names else ''
    seller.brand = brand_names[0] if brand_names else ''
    seller.model = ''

    seller.all_categories = not bool(category_names) or 'Multi' in category_names
    seller.all_countries = not bool(country_names) or 'Multi' in country_names
    seller.all_brands = not bool(brand_names) or 'Multi' in brand_names
    seller.all_models = True

    seller.save()

    seller.selected_categories.clear()
    seller.selected_countries.clear()
    seller.selected_brands.clear()
    seller.selected_models.clear()

    for category_name in category_names:
        if category_name == 'Multi':
            continue

        category, _ = PartCategory.objects.get_or_create(name=category_name)
        seller.selected_categories.add(category)

    for country_name in country_names:
        if country_name == 'Multi':
            continue

        country, _ = Country.objects.get_or_create(name=country_name)
        seller.selected_countries.add(country)

    for brand_name in brand_names:
        if brand_name == 'Multi':
            continue

        matches = Brand.objects.filter(
            name__iexact=brand_name,
            transport_type=row['transport_type']
        )

        for brand in matches:
            seller.selected_brands.add(brand)

    return created


@admin.register(Country)
class CountryAdmin(admin.ModelAdmin):
    list_display = ('id', 'name')
    search_fields = ('name',)


@admin.register(Brand)
class BrandAdmin(admin.ModelAdmin):
    list_display = ('id', 'name', 'country', 'transport_type')
    list_filter = ('country', 'transport_type')
    search_fields = ('name',)


@admin.register(CarModel)
class CarModelAdmin(admin.ModelAdmin):
    list_display = ('id', 'name', 'brand', 'transport_type')
    list_filter = ('transport_type', 'brand', 'brand__country')
    search_fields = ('name',)


@admin.register(PartCategory)
class PartCategoryAdmin(admin.ModelAdmin):
    list_display = ('id', 'name')
    search_fields = ('name',)


@admin.register(BroadcastSettings)
class BroadcastSettingsAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'mode',
        'wave_size',
        'wave_interval_minutes',
        'emergency_stop',
        'updated_at',
    )

    readonly_fields = ('updated_at',)

    fieldsets = (
        ('Broadcast Control', {
            'fields': (
                'mode',
                'wave_size',
                'wave_interval_minutes',
                'emergency_stop',
                'updated_at',
            )
        }),
    )

    def has_add_permission(self, request):
        if BroadcastSettings.objects.exists():
            return False
        return super().has_add_permission(request)

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(Request)
class RequestAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'transport_type',
        'country',
        'brand',
        'model',
        'category',
        'city',
        'phone',
        'status',
        'created_at'
    )

    list_filter = (
        'transport_type',
        'status',
        'city',
        'category'
    )

    search_fields = (
        'country',
        'brand',
        'model',
        'article',
        'description',
        'phone'
    )

    readonly_fields = ('buyer_contact',)


BUYER_INLINE_AGGREGATE_READONLY = (
    'first_seen_at',
    'last_seen_at',
    'requests_count',
    'created_at',
    'updated_at',
)


class BuyerVehicleInline(admin.TabularInline):
    model = BuyerVehicle
    extra = 0
    readonly_fields = BUYER_INLINE_AGGREGATE_READONLY + (
        'brand_normalized',
        'model_normalized',
    )
    fields = (
        'transport_type',
        'brand',
        'model',
        'brand_normalized',
        'model_normalized',
        'first_seen_at',
        'last_seen_at',
        'requests_count',
    )


class BuyerCategoryInterestInline(admin.TabularInline):
    model = BuyerCategoryInterest
    extra = 0
    readonly_fields = BUYER_INLINE_AGGREGATE_READONLY + ('category_normalized',)
    fields = (
        'category',
        'category_normalized',
        'first_seen_at',
        'last_seen_at',
        'requests_count',
    )


class BuyerCityInterestInline(admin.TabularInline):
    model = BuyerCityInterest
    extra = 0
    readonly_fields = BUYER_INLINE_AGGREGATE_READONLY + ('city_normalized',)
    fields = (
        'city',
        'city_normalized',
        'interest_type',
        'first_seen_at',
        'last_seen_at',
        'requests_count',
    )


class ContactConsentInline(admin.TabularInline):
    model = ContactConsent
    extra = 0
    readonly_fields = ('created_at', 'updated_at')
    fields = (
        'channel',
        'purpose',
        'status',
        'source',
        'consent_text_version',
        'evidence_reference',
        'consented_at',
        'revoked_at',
    )


@admin.action(description='Отметить как тестовые контакты')
def mark_buyer_contacts_as_test(modeladmin, request, queryset):
    updated = queryset.update(is_test_contact=True)
    messages.success(request, f'Отмечено как тестовые контакты: {updated}.')


@admin.action(description='Снять признак тестового контакта')
def unmark_buyer_contacts_as_test(modeladmin, request, queryset):
    updated = queryset.update(is_test_contact=False)
    messages.success(request, f'Снят признак тестового контакта: {updated}.')


@admin.register(BuyerContact)
class BuyerContactAdmin(admin.ModelAdmin):
    list_display = (
        'masked_phone',
        'primary_city',
        'vehicles_summary',
        'categories_summary',
        'requests_count',
        'last_request_at',
        'last_search_scope',
        'marketing_consent_status',
        'is_test_contact',
        'status',
    )
    list_filter = (
        'is_test_contact',
        'status',
        BuyerMarketingConsentFilter,
        BuyerActivityFilter,
        BuyerRequestCountFilter,
        'primary_country',
        BuyerPrimaryCityFilter,
        BuyerTransportTypeFilter,
        BuyerBrandFilter,
        BuyerModelFilter,
        BuyerCategoryFilter,
        'last_search_scope',
    )
    search_fields = (
        'phone_normalized',
        'primary_country',
        'primary_city',
        'vehicles__brand',
        'vehicles__model',
        'category_interests__category',
    )
    ordering = ('-last_request_at', '-id')
    list_select_related = ('portal_access',)
    readonly_fields = (
        'phone_normalized',
        'primary_country',
        'primary_city',
        'first_request_at',
        'last_request_at',
        'requests_count',
        'last_search_scope',
        'city_scope_requests_count',
        'kazakhstan_scope_requests_count',
        'custom_scope_requests_count',
        'portal_access',
        'created_at',
        'updated_at',
    )
    fields = (
        'phone_normalized',
        'status',
        'is_test_contact',
        'primary_country',
        'primary_city',
        'first_request_at',
        'last_request_at',
        'requests_count',
        'last_search_scope',
        'city_scope_requests_count',
        'kazakhstan_scope_requests_count',
        'custom_scope_requests_count',
        'portal_access',
        'source',
        'created_at',
        'updated_at',
    )
    inlines = (
        BuyerVehicleInline,
        BuyerCategoryInterestInline,
        BuyerCityInterestInline,
        ContactConsentInline,
    )
    actions = (
        mark_buyer_contacts_as_test,
        unmark_buyer_contacts_as_test,
    )

    def get_queryset(self, request):
        marketing_consents = ContactConsent.objects.filter(
            channel=CONTACT_CONSENT_CHANNEL_WHATSAPP,
            purpose=CONTACT_CONSENT_PURPOSE_MARKETING,
        )
        return super().get_queryset(request).select_related(
            'portal_access',
        ).prefetch_related(
            Prefetch(
                'vehicles',
                queryset=BuyerVehicle.objects.order_by('-last_seen_at', '-id'),
            ),
            Prefetch(
                'category_interests',
                queryset=BuyerCategoryInterest.objects.order_by(
                    '-last_seen_at',
                    '-id',
                ),
            ),
            Prefetch(
                'city_interests',
                queryset=BuyerCityInterest.objects.order_by('-last_seen_at', '-id'),
            ),
            Prefetch(
                'consents',
                queryset=marketing_consents,
                to_attr='marketing_consents',
            ),
        )

    def get_search_results(self, request, queryset, search_term):
        queryset, may_have_duplicates = super().get_search_results(
            request,
            queryset,
            search_term,
        )
        return queryset.distinct(), True

    @admin.display(description='Телефон', ordering='phone_normalized')
    def masked_phone(self, obj):
        return mask_phone(obj.phone_normalized)

    @admin.display(description='Автомобили')
    def vehicles_summary(self, obj):
        prefetched = getattr(obj, '_prefetched_objects_cache', {})
        vehicles = prefetched.get('vehicles')
        if vehicles is None:
            vehicles = obj.vehicles.all()
        return build_vehicle_summary(vehicles)

    @admin.display(description='Категории')
    def categories_summary(self, obj):
        prefetched = getattr(obj, '_prefetched_objects_cache', {})
        interests = prefetched.get('category_interests')
        if interests is None:
            interests = obj.category_interests.all()
        return build_category_summary(interests)

    @admin.display(description='Рекламное согласие')
    def marketing_consent_status(self, obj):
        consents = getattr(obj, 'marketing_consents', None)
        if consents:
            return marketing_consent_label(consents[0].status)
        return marketing_consent_label(None)


@admin.register(BuyerVehicle)
class BuyerVehicleAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'buyer',
        'transport_type',
        'brand',
        'model',
        'requests_count',
        'last_seen_at',
    )
    list_filter = ('transport_type',)
    search_fields = (
        'brand',
        'model',
        'brand_normalized',
        'model_normalized',
        'buyer__phone_normalized',
    )
    autocomplete_fields = ('buyer',)
    readonly_fields = (
        'brand_normalized',
        'model_normalized',
        'created_at',
        'updated_at',
    )

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('buyer')


@admin.register(BuyerCategoryInterest)
class BuyerCategoryInterestAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'buyer',
        'category',
        'requests_count',
        'last_seen_at',
    )
    search_fields = ('category', 'category_normalized', 'buyer__phone_normalized')
    autocomplete_fields = ('buyer',)
    readonly_fields = ('category_normalized', 'created_at', 'updated_at')

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('buyer')


@admin.register(BuyerCityInterest)
class BuyerCityInterestAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'buyer',
        'city',
        'interest_type',
        'requests_count',
        'last_seen_at',
    )
    list_filter = ('interest_type',)
    search_fields = ('city', 'city_normalized', 'buyer__phone_normalized')
    autocomplete_fields = ('buyer',)
    readonly_fields = ('city_normalized', 'created_at', 'updated_at')

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('buyer')


@admin.register(ContactConsent)
class ContactConsentAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'buyer',
        'channel',
        'purpose',
        'status',
        'source',
        'consented_at',
        'revoked_at',
    )
    list_filter = ('channel', 'purpose', 'status', 'source')
    search_fields = (
        'buyer__phone_normalized',
        'consent_text_version',
        'evidence_reference',
    )
    autocomplete_fields = ('buyer',)
    readonly_fields = ('created_at', 'updated_at')

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('buyer')


@admin.register(BuyerAudience)
class BuyerAudienceAdmin(admin.ModelAdmin):
    form = BuyerAudienceAdminForm
    list_display = (
        'name',
        'is_active',
        'criteria_summary',
        'preview_link',
        'updated_at',
    )
    list_filter = ('is_active', 'updated_at')
    search_fields = ('name', 'description')
    readonly_fields = (
        'created_at',
        'updated_at',
        'criteria_readonly_summary',
    )
    fields = (
        'name',
        'description',
        'is_active',
        'countries',
        'cities',
        'transport_types',
        'brands',
        'models',
        'categories',
        'search_scopes',
        'activity_period',
        'request_count_min',
        'request_count_max',
        'criteria_readonly_summary',
        'created_at',
        'updated_at',
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                '<path:object_id>/preview/',
                self.admin_site.admin_view(self.preview_view),
                name='core_buyeraudience_preview',
            ),
        ]
        return custom_urls + urls

    def preview_view(self, request, object_id):
        audience = get_object_or_404(BuyerAudience, pk=object_id)
        preview = preview_buyer_audience(audience)
        context = {
            **self.admin_site.each_context(request),
            'opts': self.model._meta,
            'audience': audience,
            'preview': preview,
            'sample_contacts': preview.sample_contacts,
            'criteria_rows': format_criteria_details(audience.criteria),
            'no_criteria': not audience_criteria_has_filters(audience.criteria),
            'calculated_at': timezone.now(),
            'title': f'Предпросмотр: {audience.name}',
        }
        return render(request, 'admin/core/buyeraudience/preview.html', context)

    @admin.display(description='Критерии')
    def criteria_summary(self, obj):
        return format_criteria_summary(obj.criteria)

    @admin.display(description='Критерии')
    def criteria_readonly_summary(self, obj):
        rows = format_criteria_details(obj.criteria)
        return mark_safe('<br>'.join(f'<strong>{label}:</strong> {escape(value)}' for label, value in rows))

    @admin.display(description='Предпросмотр')
    def preview_link(self, obj):
        if not obj.pk:
            return '—'
        url = reverse('admin:core_buyeraudience_preview', args=[obj.pk])
        return mark_safe(f'<a href="{escape(url)}">Предпросмотр</a>')


@admin.action(description='Включить получение заявок')
def enable_receive_requests(modeladmin, request, queryset):
    updated = queryset.update(receive_requests=True)
    messages.success(request, f'Получение заявок включено: {updated} продавцов.')


@admin.action(description='Выключить получение заявок')
def disable_receive_requests(modeladmin, request, queryset):
    updated = queryset.update(receive_requests=False)
    messages.warning(request, f'Получение заявок выключено: {updated} продавцов.')


@admin.action(description='Пометить как тестовых продавцов')
def mark_as_test_seller(modeladmin, request, queryset):
    updated = queryset.update(is_test_seller=True)
    messages.success(request, f'Помечено как тестовые продавцы: {updated}.')


@admin.action(description='Снять признак тестовых продавцов')
def unmark_as_test_seller(modeladmin, request, queryset):
    updated = queryset.update(is_test_seller=False)
    messages.warning(request, f'Снят признак тестовых продавцов: {updated}.')


@admin.register(Seller)
class SellerAdmin(admin.ModelAdmin):
    change_list_template = 'admin/core/seller/change_list.html'

    list_display = (
        'id',
        'name',
        'whatsapp',
        'phone2',
        'transport_type',
        'seller_type',
        'dispatch_priority',
        'receive_requests',
        'is_test_seller',
        'city',
        'market_location',
        'is_active',
        'is_paused',
    )

    list_editable = (
        'dispatch_priority',
        'receive_requests',
        'is_test_seller',
    )

    list_filter = (
        'transport_type',
        'seller_type',
        'receive_requests',
        'is_test_seller',
        'is_active',
        'is_paused',
        'all_categories',
        'all_brands',
        'all_models',
        'all_countries',
        'city'
    )

    search_fields = (
        'name',
        'whatsapp',
        'phone2',
        'brand',
        'model',
        'market_location',
        'notes',
    )

    filter_horizontal = (
        'selected_categories',
        'selected_countries',
        'selected_brands',
        'selected_models'
    )

    actions = (
        enable_receive_requests,
        disable_receive_requests,
        mark_as_test_seller,
        unmark_as_test_seller,
    )

    fieldsets = (
        ('Основное', {
            'fields': (
                'name',
                'whatsapp',
                'phone2',
                'seller_type',
                'transport_type',
                'city',
                'market_location',
                'dispatch_priority',
                'is_active',
                'is_paused',
                'receive_requests',
                'is_test_seller',
                'notes',
            )
        }),

        ('Старый одиночный режим', {
            'fields': (
                'category',
                'country_fk',
                'brand_fk',
                'model_fk'
            )
        }),

        ('Новый множественный выбор', {
            'fields': (
                'selected_categories',
                'selected_countries',
                'selected_brands',
                'selected_models'
            )
        }),

        ('Режимы "Все"', {
            'fields': (
                'all_categories',
                'all_countries',
                'all_brands',
                'all_models'
            )
        }),
    )

    def get_urls(self):
        urls = super().get_urls()

        custom_urls = [
            path(
                'import-xlsx/',
                self.admin_site.admin_view(self.import_xlsx_view),
                name='core_seller_import_xlsx'
            ),
        ]

        return custom_urls + urls

    def import_xlsx_view(self, request):
        context = {
            **self.admin_site.each_context(request),
            'title': 'Импорт продавцов XLSX',
            'opts': self.model._meta,
            'form': SellerImportForm(),
            'preview_rows': None,
            'errors': [],
        }

        if request.method == 'POST' and request.POST.get('action') == 'preview':
            form = SellerImportForm(request.POST, request.FILES)

            if form.is_valid():
                rows, errors = parse_sellers_xlsx(request.FILES['file'])

                preview_rows = []

                for row in rows:
                    exists = find_seller_by_whatsapp(row['whatsapp']) is not None
                    preview_rows.append({
                        **row,
                        'status': 'Обновление' if exists else 'Новый',
                    })

                request.session['seller_import_rows'] = rows

                context.update({
                    'form': form,
                    'preview_rows': preview_rows,
                    'errors': errors,
                })

                return render(request, 'admin/core/seller/import_sellers.html', context)

        if request.method == 'POST' and request.POST.get('action') == 'import':
            rows = request.session.get('seller_import_rows', [])

            created_count = 0
            updated_count = 0
            failed_count = 0

            for row in rows:
                try:
                    created = import_seller_row(row)

                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                except Exception as exc:
                    failed_count += 1
                    messages.error(
                        request,
                        f"Ошибка строки {row.get('row_number')}: "
                        f"{row.get('seller_name')} — {exc}"
                    )

            request.session.pop('seller_import_rows', None)

            messages.success(
                request,
                f'Импорт завершён. Создано: {created_count}. '
                f'Обновлено: {updated_count}. Ошибок: {failed_count}.'
            )

            return redirect('..')

        return render(request, 'admin/core/seller/import_sellers.html', context)


@admin.register(Match)
class MatchAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'request',
        'seller',
        'status',
        'created_at',
        'sent_at'
    )

    list_filter = ('status',)

    search_fields = (
        'request__phone',
        'seller__name'
    )


@admin.action(description='Остановить выбранные волны')
def pause_dispatches(modeladmin, request, queryset):
    updated = queryset.exclude(status=RequestDispatch.STATUS_SENT).update(
        status=RequestDispatch.STATUS_PAUSED
    )
    messages.warning(request, f'Остановлено волн/отправок: {updated}.')


@admin.action(description='Вернуть выбранные волны в очередь')
def queue_dispatches(modeladmin, request, queryset):
    updated = queryset.exclude(status=RequestDispatch.STATUS_SENT).update(
        status=RequestDispatch.STATUS_QUEUED
    )
    messages.success(request, f'Возвращено в очередь: {updated}.')


@admin.register(RequestDispatch)
class RequestDispatchAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'request',
        'seller',
        'wave_number',
        'position_number',
        'status',
        'scheduled_at',
        'sent_at',
        'created_at'
    )

    list_filter = (
        'status',
        'wave_number',
        'scheduled_at',
        'sent_at'
    )

    search_fields = (
        'request__phone',
        'request__brand',
        'request__model',
        'seller__name',
        'seller__whatsapp'
    )

    readonly_fields = (
        'created_at',
    )

    actions = (
        pause_dispatches,
        queue_dispatches,
    )


@admin.register(Feedback)
class FeedbackAdmin(admin.ModelAdmin):
    list_display = ('id', 'name', 'phone', 'created_at')
    search_fields = ('name', 'phone', 'message')
    readonly_fields = ('created_at',)
    ordering = ('-created_at',)


@admin.register(WhatsAppMessageLog)
class WhatsAppMessageLogAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'created_at',
        'seller_name',
        'phone_clean',
        'is_success',
        'status_text',
        'message_id',
    )
    search_fields = ('seller_name', 'phone_clean', 'message_id')
    list_filter = ('is_success', 'status_text')


@admin.action(description='Одобрить выбранные публикации')
def approve_instagram_publications(modeladmin, request, queryset):
    updated = 0
    for publication in queryset:
        approve_instagram_publication(publication)
        updated += 1
    modeladmin.message_user(request, f'Одобрено публикаций: {updated}')


@admin.action(description='Опубликовать выбранные карточки')
def publish_instagram_publications(modeladmin, request, queryset):
    queued = 0
    skipped = 0
    for publication in queryset:
        if publication.status == InstagramPublication.STATUS_PUBLISHED:
            skipped += 1
            continue
        queue_instagram_publication_for_processing(publication)
        queued += 1
    modeladmin.message_user(
        request,
        f'Публикация поставлена в очередь: {queued}. Пропущено (уже опубликовано): {skipped}.',
    )


@admin.action(description='Повторить публикацию для выбранных')
def retry_instagram_publications(modeladmin, request, queryset):
    queued = 0
    for publication in queryset.exclude(status=InstagramPublication.STATUS_PUBLISHED):
        queue_instagram_publication_for_processing(publication)
        queued += 1
    modeladmin.message_user(request, f'Публикация поставлена в очередь: {queued}.')


@admin.action(description='Пометить зависшие публикации как ошибку (>5 мин)')
def mark_stuck_instagram_publications_failed(modeladmin, request, queryset):
    updated = 0
    skipped = 0
    for publication in queryset.filter(status=InstagramPublication.STATUS_PUBLISHING):
        before = publication.status
        mark_stuck_instagram_publication_failed(publication)
        publication.refresh_from_db()
        if publication.status == InstagramPublication.STATUS_FAILED and before != publication.status:
            updated += 1
        else:
            skipped += 1
    modeladmin.message_user(
        request,
        f'Помечено как ошибка: {updated}. Пропущено (ещё не зависли): {skipped}.',
    )


@admin.action(description='Отменить выбранные публикации')
def cancel_instagram_publications(modeladmin, request, queryset):
    cancelled = 0
    for publication in queryset:
        if publication.status != InstagramPublication.STATUS_PUBLISHED:
            cancel_instagram_publication(publication)
            cancelled += 1
    modeladmin.message_user(request, f'Отменено публикаций: {cancelled}')


@admin.register(InstagramPublication)
class InstagramPublicationAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'request',
        'status',
        'image_preview_list',
        'created_at',
        'published_at',
    )
    list_filter = ('status', 'created_at', 'published_at')
    search_fields = (
        'request__id',
        'request__brand',
        'request__model',
        'caption',
        'instagram_media_id',
        'error_message',
    )
    readonly_fields = (
        'request',
        'image',
        'image_preview',
        'caption',
        'instagram_container_id',
        'instagram_media_id',
        'created_at',
        'publishing_started_at',
        'published_at',
    )
    fields = (
        'request',
        'status',
        'image_preview',
        'image',
        'caption',
        'instagram_container_id',
        'instagram_media_id',
        'error_message',
        'created_at',
        'publishing_started_at',
        'published_at',
    )
    actions = (
        approve_instagram_publications,
        publish_instagram_publications,
        retry_instagram_publications,
        mark_stuck_instagram_publications_failed,
        cancel_instagram_publications,
    )

    @admin.display(description='Превью')
    def image_preview_list(self, obj):
        return self._render_preview(obj, max_height=48)

    @admin.display(description='Превью карточки')
    def image_preview(self, obj):
        return self._render_preview(obj, max_height=320)

    def _render_preview(self, obj, *, max_height: int):
        from django.utils.html import format_html

        if not obj.image:
            return '—'
        return format_html(
            '<img src="{}" alt="Instagram story preview" '
            'style="max-height:{}px;border-radius:8px;border:1px solid #e5e7eb;" />',
            obj.image.url,
            max_height,
        )


def _seller_lead_external_link(url: str, label: str):
    from django.utils.html import format_html

    if not url:
        return '—'
    return format_html(
        '<a href="{}" target="_blank" rel="noopener noreferrer">{}</a>',
        url,
        label,
    )


@admin.action(description='Пометить как проверенные')
def mark_seller_leads_verified(modeladmin, request, queryset):
    from django.utils import timezone

    now = timezone.now()
    updated = 0
    for lead in queryset:
        lead.status = SellerLead.STATUS_VERIFIED
        if not lead.checked_at:
            lead.checked_at = now
        lead.save(update_fields=['status', 'checked_at', 'updated_at'])
        updated += 1
    messages.success(request, f'Помечено как проверенные: {updated}.')


@admin.action(description='Пометить как дубликаты')
def mark_seller_leads_duplicate(modeladmin, request, queryset):
    updated = queryset.update(status=SellerLead.STATUS_DUPLICATE)
    messages.warning(request, f'Помечено как дубликаты: {updated}.')


@admin.action(description='Пометить как не продавцов')
def mark_seller_leads_not_seller(modeladmin, request, queryset):
    updated = queryset.update(status=SellerLead.STATUS_NOT_SELLER)
    messages.warning(request, f'Помечено как не продавцы: {updated}.')


@admin.action(description='Пометить как «Нет WhatsApp»')
def mark_seller_leads_no_whatsapp(modeladmin, request, queryset):
    updated = queryset.update(status=SellerLead.STATUS_NO_WHATSAPP)
    messages.warning(request, f'Помечено как «Нет WhatsApp»: {updated}.')


@admin.action(description='Пометить как «Написали»')
def mark_seller_leads_contacted(modeladmin, request, queryset):
    updated = queryset.update(status=SellerLead.STATUS_CONTACTED)
    messages.success(request, f'Помечено как «Написали»: {updated}.')


class SellerLeadContactCandidateInline(admin.TabularInline):
    model = SellerLeadContactCandidate
    extra = 0
    fields = (
        'value',
        'contact_type',
        'role',
        'label',
        'confidence',
        'status',
        'is_primary',
        'source_type',
        'source_url',
        'found_at',
        'reviewed_at',
    )
    readonly_fields = ('found_at', 'reviewed_at')
    show_change_link = True


@admin.action(description='Подтвердить кандидата как основной контакт')
def approve_contact_candidates_as_primary(modeladmin, request, queryset):
    lead_ids = set(queryset.values_list('seller_lead_id', flat=True))
    if len(lead_ids) > 1:
        messages.error(
            request,
            'Нельзя подтвердить кандидатов из разных SellerLead одновременно. Выберите кандидатов одного лида.',
        )
        return
    if queryset.count() > 1:
        messages.error(
            request,
            'Нельзя подтвердить несколько кандидатов одновременно. Выберите одного кандидата для основного контакта.',
        )
        return
    candidate = queryset.first()
    if candidate is None:
        return
    try:
        candidate.approve_as_primary()
    except ValueError as exc:
        messages.error(request, str(exc))
        return
    messages.success(request, f'Основной контакт подтверждён: {candidate.value}.')


@admin.action(description='Пометить кандидата как отклонённый')
def reject_contact_candidates(modeladmin, request, queryset):
    updated = queryset.update(
        status=SellerLeadContactCandidate.STATUS_REJECTED,
        is_primary=False,
        reviewed_at=timezone.now(),
    )
    messages.warning(request, f'Отклонено кандидатов: {updated}.')


@admin.action(description='Пометить кандидата как конфликтующий')
def mark_contact_candidates_conflict(modeladmin, request, queryset):
    updated = queryset.update(
        status=SellerLeadContactCandidate.STATUS_CONFLICT,
        is_primary=False,
        reviewed_at=timezone.now(),
    )
    messages.warning(request, f'Помечено как конфликт: {updated}.')


@admin.action(description='Вернуть кандидата в ожидание проверки')
def mark_contact_candidates_pending(modeladmin, request, queryset):
    updated = queryset.update(
        status=SellerLeadContactCandidate.STATUS_PENDING,
        is_primary=False,
        reviewed_at=None,
    )
    messages.info(request, f'Возвращено в ожидание проверки: {updated}.')


@admin.register(SellerLeadContactCandidate)
class SellerLeadContactCandidateAdmin(admin.ModelAdmin):
    list_display = (
        'seller_lead',
        'value',
        'contact_type',
        'role',
        'label',
        'confidence',
        'status',
        'is_primary',
        'source_type',
        'found_at',
        'reviewed_at',
    )
    list_filter = (
        'contact_type',
        'role',
        'confidence',
        'status',
        'is_primary',
        'source_type',
        'found_at',
    )
    search_fields = (
        'seller_lead__name',
        'seller_lead__instagram_username',
        'value',
        'label',
        'notes',
    )
    readonly_fields = (
        'created_at',
        'updated_at',
        'whatsapp_link',
        'source_link',
    )
    actions = (
        approve_contact_candidates_as_primary,
        reject_contact_candidates,
        mark_contact_candidates_conflict,
        mark_contact_candidates_pending,
    )
    fieldsets = (
        ('Контакт', {
            'fields': (
                'seller_lead',
                'contact_type',
                'value',
                'role',
                'label',
                'confidence',
                'status',
                'is_primary',
                'whatsapp_link',
            ),
        }),
        ('Источник', {
            'fields': (
                'source_type',
                'source_url',
                'source_link',
                'source_text',
                'found_at',
                'reviewed_at',
            ),
        }),
        ('Служебное', {
            'fields': (
                'notes',
                'created_at',
                'updated_at',
            ),
        }),
    )

    @admin.display(description='WhatsApp')
    def whatsapp_link(self, obj):
        return _seller_lead_external_link(
            obj.get_whatsapp_url(),
            'Открыть WhatsApp',
        )

    @admin.display(description='Источник')
    def source_link(self, obj):
        return _seller_lead_external_link(obj.source_url, 'Открыть источник')


@admin.register(SellerLead)
class SellerLeadAdmin(admin.ModelAdmin):
    inlines = (SellerLeadContactCandidateInline,)
    list_display = (
        'name',
        'instagram_username',
        'whatsapp',
        'city',
        'category',
        'status',
        'source_type',
        'collected_at',
        'checked_at',
    )
    list_filter = (
        'status',
        'source_type',
        'city',
        'category',
        'collected_at',
        'checked_at',
    )
    search_fields = (
        'name',
        'instagram_username',
        'whatsapp',
        'city',
        'category',
        'car_brands',
        'notes',
        'profile_description',
    )
    readonly_fields = (
        'created_at',
        'updated_at',
        'instagram_profile_link',
        'whatsapp_link',
        'website_link',
        'source_link',
    )
    actions = (
        mark_seller_leads_verified,
        mark_seller_leads_duplicate,
        mark_seller_leads_not_seller,
        mark_seller_leads_no_whatsapp,
        mark_seller_leads_contacted,
    )
    fieldsets = (
        ('Основное', {
            'fields': (
                'name',
                'status',
                'source_type',
                'collected_at',
                'checked_at',
            ),
        }),
        ('Контакты', {
            'fields': (
                'instagram_username',
                'instagram_url',
                'instagram_profile_link',
                'whatsapp',
                'whatsapp_source_url',
                'whatsapp_source_text',
                'whatsapp_confidence',
                'whatsapp_found_at',
                'whatsapp_link',
                'website_url',
                'website_link',
            ),
        }),
        ('Профиль', {
            'fields': (
                'city',
                'category',
                'car_brands',
                'profile_description',
            ),
        }),
        ('Источник', {
            'fields': (
                'source_url',
                'source_link',
                'notes',
            ),
        }),
        ('Служебное', {
            'fields': (
                'created_at',
                'updated_at',
            ),
        }),
    )

    @admin.display(description='Instagram')
    def instagram_profile_link(self, obj):
        return _seller_lead_external_link(
            obj.get_instagram_profile_url(),
            'Открыть Instagram',
        )

    @admin.display(description='WhatsApp')
    def whatsapp_link(self, obj):
        return _seller_lead_external_link(
            obj.get_whatsapp_url(),
            'Открыть WhatsApp',
        )

    @admin.display(description='Сайт')
    def website_link(self, obj):
        return _seller_lead_external_link(obj.website_url, 'Открыть сайт')

    @admin.display(description='Источник')
    def source_link(self, obj):
        return _seller_lead_external_link(obj.source_url, 'Открыть источник')


def _format_json_for_admin(value) -> str:
    import json

    if not value:
        return '—'
    try:
        rendered = json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True)
    except TypeError:
        rendered = str(value)
    return mark_safe(f'<pre style="white-space:pre-wrap;">{escape(rendered)}</pre>')


@admin.register(SellerLeadPipelineRun)
class SellerLeadPipelineRunAdmin(admin.ModelAdmin):
    list_display = (
        'started_at',
        'trigger',
        'status',
        'city',
        'search_term',
        'category',
        'lead_limit',
        'discovery_new_profiles',
        'enrichment_saved_contacts',
        'enrichment_conflicts',
    )
    list_filter = ('status', 'trigger', 'city', 'category', 'rotation_enabled')
    search_fields = ('run_uuid',)
    ordering = ('-started_at',)
    readonly_fields = (
        'run_uuid',
        'trigger',
        'status',
        'is_dry_run',
        'city',
        'category',
        'search_term',
        'rotation_enabled',
        'rotation_slug',
        'rotation_index',
        'search_limit',
        'lead_limit',
        'max_queries_per_lead',
        'skip_discovery',
        'skip_enrichment',
        'cooldown_minutes',
        'force_run',
        'started_at',
        'finished_at',
        'discovery_stats_display',
        'enrichment_stats_display',
        'created_lead_ids_display',
        'error_message',
        'skip_reason',
    )

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    @admin.display(description='Новых профилей')
    def discovery_new_profiles(self, obj):
        return int((obj.discovery_stats or {}).get('new_profiles', 0))

    @admin.display(description='Сохранено WhatsApp')
    def enrichment_saved_contacts(self, obj):
        return int((obj.enrichment_stats or {}).get('saved_primary_whatsapp', 0))

    @admin.display(description='Конфликтов')
    def enrichment_conflicts(self, obj):
        return int((obj.enrichment_stats or {}).get('conflicts', 0))

    @admin.display(description='Discovery stats')
    def discovery_stats_display(self, obj):
        return _format_json_for_admin(obj.discovery_stats)

    @admin.display(description='Enrichment stats')
    def enrichment_stats_display(self, obj):
        return _format_json_for_admin(obj.enrichment_stats)

    @admin.display(description='Созданные лиды (ID)')
    def created_lead_ids_display(self, obj):
        return _format_json_for_admin(obj.created_lead_ids)