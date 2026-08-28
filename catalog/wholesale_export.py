"""Dynamic public wholesale price list (XLSX)."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from catalog.wholesale import (
    VAT_INCLUDED,
    VAT_EXCLUDED,
    build_wholesale_terms_snapshot,
    public_stock_xlsx_value,
    public_wholesale_unit_price,
    wholesale_fitment_text,
    wholesale_products_qs,
    wholesale_storefront_condition_lines,
)

XLSX_CONTENT_TYPE = (
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)
AVAILABILITY_TEXT = 'Уточняется при заказе'


def _vat_wholesale_header(snapshot):
    vat_mode = (snapshot or {}).get('vat_mode')
    if vat_mode == VAT_INCLUDED:
        return 'Оптовая цена, ₸ с НДС'
    if vat_mode == VAT_EXCLUDED:
        return 'Оптовая цена, ₸ без НДС'
    return 'Оптовая цена, ₸'


def _price_headers(snapshot):
    return [
        'Артикул',
        'Наименование',
        'Марка',
        'Модель / применяемость',
        'Розничная цена, ₸',
        _vat_wholesale_header(snapshot),
        'Минимальный общий заказ',
        'Наличие',
        'Продавец',
    ]


def build_wholesale_price_workbook(seller):
    snapshot = build_wholesale_terms_snapshot(seller)
    workbook = Workbook()
    prices = workbook.active
    prices.title = 'Прайс'
    headers = _price_headers(snapshot)
    prices.append(headers)
    header_font = Font(bold=True)
    for cell in prices[1]:
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical='center')

    min_qty = int(snapshot.get('min_order_qty') or 0)
    seller_name = getattr(seller, 'name', '') or ''
    products = wholesale_products_qs(seller).order_by('article', 'id')
    for product in products:
        brand_name = ''
        if getattr(product, 'brand', None) and product.brand.name:
            brand_name = product.brand.name
        elif getattr(product, 'car_model', None) and product.car_model.brand:
            brand_name = product.car_model.brand.name
        wholesale_price = public_wholesale_unit_price(product)
        prices.append([
            product.article or '',
            product.title or '',
            brand_name,
            wholesale_fitment_text(product),
            int(product.price) if product.price is not None else '',
            int(wholesale_price) if wholesale_price is not None else '',
            min_qty or '',
            public_stock_xlsx_value(product),
            seller_name,
        ])

    for index, width in enumerate((18, 42, 16, 42, 18, 22, 22, 28, 22), start=1):
        prices.column_dimensions[get_column_letter(index)].width = width

    terms_sheet = workbook.create_sheet('Условия')
    terms_sheet['A1'] = 'Условия оптовой покупки'
    terms_sheet['A1'].font = Font(bold=True, size=14)
    terms_sheet['A2'] = seller_name
    row = 4
    for line in wholesale_storefront_condition_lines(snapshot):
        terms_sheet.cell(row=row, column=1, value=line)
        row += 1
    terms_sheet.column_dimensions['A'].width = 72
    return workbook


def wholesale_price_xlsx_bytes(seller):
    workbook = build_wholesale_price_workbook(seller)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
