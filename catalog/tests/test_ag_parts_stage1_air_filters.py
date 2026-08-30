from importlib import import_module

from django.apps import apps
from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from io import BytesIO
from openpyxl import load_workbook

from catalog.ag_parts_air_filters import (
    EXCLUDED_ARTICLES,
    LEGACY_NULL_DUPLICATES,
    ORIGINAL_WHOLESALE_ARTICLES,
    STAGE1_AIR_FILTERS,
)
from catalog.models import Product, ProductPriceTier, SellerProfile, SellerWholesaleTerms
from catalog.wholesale import has_public_wholesale_offer, wholesale_products_qs
from marketing.models import MarketingWhatsAppTemplate
from orders.models import Order, OrderItem


STAGE1 = import_module('catalog.migrations.0027_ag_parts_stage1_air_filters')

KNOWN_TITLES = {
    '1109101XGW01A': 'Воздушный фильтр Haval Dargo',
    '1109104XGW02A': 'Воздушный фильтр Haval Jolion',
    '1109130U2400': 'Воздушный фильтр JAC JS6',
    '1109190CR01': 'Воздушный фильтр Changan UNI-K',
    '151000025AA': 'Воздушный фильтр Exeed TXL',
    '151000079AA': 'Воздушный фильтр Chery Tiggo 7 Pro',
    '151000187AA': 'Воздушный фильтр Exeed TXL',
    '6600131687': 'Воздушный фильтр Geely Coolray',
    'S1010140400': 'Воздушный фильтр Changan CS35',
    'S3010140903': 'Воздушный фильтр Changan UNI-V',
    'T151109111': 'Воздушный фильтр Chery Tiggo 7',
    'X01-90000014': 'Воздушный фильтр Li Auto L7',
}

EXISTING_ACTIVE = (
    '1109190CR01',
    '151000025AA',
    'S1010140400',
    'T151109111',
)
EXISTING_HIDDEN = (
    '1109101XGW01A',
    '1109104XGW02A',
    '1109130U2400',
    '151000079AA',
    '151000187AA',
    '6600131687',
    'S3010140903',
    'X01-90000014',
)
TO_CREATE = (
    '1109110XP6EXACHS',
    '1109110XP64XA',
    '1109120U8710',
    '1109140W5000',
    '151000151AA',
    '2032047000',
    'FAE1109160',
    'J691109111',
    'M111109111',
)


def _make_seller(username, name, phone, **kwargs):
    user = User.objects.create_user(username=username, password='secret12345')
    defaults = {
        'user': user,
        'name': name,
        'phone': phone,
        'city': 'Алматы',
        'address': 'г. Алматы, ул. Тестовая, 1',
        'wholesale_enabled': True,
        'wholesale_min_order_qty': 10,
    }
    defaults.update(kwargs)
    return SellerProfile.objects.create(**defaults)


class AgPartsStage1AirFiltersTests(TestCase):
    def setUp(self):
        self.seller = _make_seller('ag-parts-air', 'AG Parts', '77771360740')
        self.assertEqual(self.seller.slug, 'ag-parts')
        self.terms = SellerWholesaleTerms.objects.create(
            seller=self.seller,
            vat_mode=SellerWholesaleTerms.VAT_INCLUDED,
            prepayment_percent=100,
            pickup_city='Алматы',
            primary_carrier='DPD Kazakhstan',
        )
        self.whitelist_products = []
        for index, article in enumerate(sorted(ORIGINAL_WHOLESALE_ARTICLES)):
            product = self._product(
                title=f'Approved {article}',
                article=article,
                slug=f'ag-orig-{index}',
                price=2000 + index,
                stock_qty=11 + index,
                publish_to_sellers=True,
                publish_to_kaspi=(index == 0),
            )
            ProductPriceTier.objects.create(product=product, min_qty=1, price=400 + index)
            self.whitelist_products.append(product)

        for article in EXISTING_ACTIVE:
            retail, _wholesale = STAGE1_AIR_FILTERS[article]
            self._product(
                title=KNOWN_TITLES[article],
                article=article,
                slug=f'ag-air-act-{article.lower()}',
                price=retail - 50,
                stock_qty=17 if article == '1109190CR01' else None,
                status='active',
                publish_to_sellers=False,
                publish_to_kaspi=(article == 'T151109111'),
            )
        for article in EXISTING_HIDDEN:
            retail, _wholesale = STAGE1_AIR_FILTERS[article]
            self._product(
                title=KNOWN_TITLES[article],
                article=article,
                slug=f'ag-air-hid-{article.lower()}',
                price=retail,
                status='hidden',
                publish_to_sellers=False,
            )

        self.legacy_cr = Product.objects.create(
            title='Legacy CR',
            price=2750,
            seller_name='AG Parts',
            seller_profile=None,
            whatsapp_number='+77771360740',
            status='hidden',
            publish_to_sellers=False,
            city='Алматы',
            article='1109190CR01',
            slug='ag-legacy-cr',
        )
        self.legacy_txl = Product.objects.create(
            title='Legacy TXL',
            price=1150,
            seller_name='AG Parts',
            seller_profile=None,
            whatsapp_number='+77771360740',
            status='hidden',
            publish_to_sellers=False,
            city='Алматы',
            article='151000025AA',
            slug='ag-legacy-txl',
        )
        self.legacy_cs35 = Product.objects.create(
            title='Legacy CS35',
            price=1680,
            seller_name='AG Parts',
            seller_profile=None,
            whatsapp_number='+77771360740',
            status='hidden',
            publish_to_sellers=False,
            city='Алматы',
            article='S1010140400',
            slug='ag-legacy-cs35',
        )

        for article in sorted(EXCLUDED_ARTICLES):
            self._product(
                title=f'Excluded {article}',
                article=article,
                slug=f'ag-excl-{article.lower()}',
                price=1800,
                status='active',
                publish_to_sellers=False,
            )

        self.snapshot = self._snapshot_original()
        self.terms_snapshot = {
            'vat_mode': self.terms.vat_mode,
            'prepayment_percent': self.terms.prepayment_percent,
            'pickup_city': self.terms.pickup_city,
            'primary_carrier': self.terms.primary_carrier,
        }
        self.template_snapshot = self._template_snapshot()

    def _product(self, **kwargs):
        defaults = {
            'seller_name': self.seller.name,
            'seller_profile': self.seller,
            'whatsapp_number': '+77771360740',
            'status': 'active',
            'city': 'Алматы',
            'publish_to_sellers': False,
            'publish_to_kaspi': False,
        }
        defaults.update(kwargs)
        return Product.objects.create(**defaults)

    def _snapshot_original(self):
        data = {}
        for product in self.whitelist_products:
            product.refresh_from_db()
            tier = product.price_tiers.get(min_qty=1, is_active=True)
            data[product.article] = {
                'price': product.price,
                'stock_qty': product.stock_qty,
                'publish_to_sellers': product.publish_to_sellers,
                'publish_to_kaspi': product.publish_to_kaspi,
                'status': product.status,
                'seller_profile_id': product.seller_profile_id,
                'tier_price': tier.price,
                'tier_id': tier.pk,
                'title': product.title,
            }
        return data

    def _template_snapshot(self):
        template = MarketingWhatsAppTemplate.objects.filter(
            meta_template_name='zpt_ag_parts_wholesale_v1',
            language_code='ru',
        ).first()
        if template is None:
            return None
        return {
            'meta_status': template.meta_status,
            'meta_template_id': template.meta_template_id,
            'body': template.body_text,
            'header': template.header_text,
        }

    def _run(self):
        STAGE1.upsert_ag_parts_stage1_air_filters(apps, None)
        STAGE1.upsert_ag_parts_stage1_air_filters(apps, None)

    def test_stage1_set_and_total_wholesale(self):
        self._run()
        wholesale = list(
            wholesale_products_qs(self.seller).order_by('article').values_list(
                'article', flat=True
            )
        )
        self.assertEqual(len(wholesale), 47)
        self.assertEqual(
            set(wholesale),
            set(ORIGINAL_WHOLESALE_ARTICLES) | set(STAGE1_AIR_FILTERS),
        )
        self.assertEqual(len(STAGE1_AIR_FILTERS), 21)
        air = [
            article for article in wholesale if article in STAGE1_AIR_FILTERS
        ]
        self.assertEqual(len(air), 21)
        self.assertEqual(set(air), set(STAGE1_AIR_FILTERS))

    def test_exact_prices_tiers_and_flags(self):
        self._run()
        for article, (retail, wholesale) in STAGE1_AIR_FILTERS.items():
            product = Product.objects.get(seller_profile=self.seller, article=article)
            self.assertEqual(product.status, 'active')
            self.assertEqual(product.seller_profile_id, self.seller.pk)
            self.assertTrue(product.publish_to_sellers)
            self.assertEqual(product.price, retail)
            tiers = list(product.price_tiers.filter(is_active=True))
            self.assertEqual(len(tiers), 1)
            self.assertEqual(tiers[0].min_qty, 1)
            self.assertEqual(tiers[0].price, wholesale)
            self.assertTrue(has_public_wholesale_offer(product, self.seller))
            if article in KNOWN_TITLES:
                self.assertEqual(product.title, KNOWN_TITLES[article])
            else:
                self.assertEqual(product.title, f'Воздушный фильтр {article}')
            if article == '1109190CR01':
                self.assertEqual(product.stock_qty, 17)
            elif article in TO_CREATE:
                self.assertIsNone(product.stock_qty)
            if article == 'T151109111':
                self.assertTrue(product.publish_to_kaspi)
            else:
                self.assertFalse(product.publish_to_kaspi)

    def test_original_26_unchanged(self):
        self._run()
        for product in self.whitelist_products:
            product.refresh_from_db()
            snap = self.snapshot[product.article]
            self.assertEqual(product.price, snap['price'])
            self.assertEqual(product.stock_qty, snap['stock_qty'])
            self.assertEqual(product.publish_to_sellers, snap['publish_to_sellers'])
            self.assertEqual(product.publish_to_kaspi, snap['publish_to_kaspi'])
            self.assertEqual(product.status, snap['status'])
            self.assertEqual(product.title, snap['title'])
            tier = product.price_tiers.get(min_qty=1, is_active=True)
            self.assertEqual(tier.price, snap['tier_price'])
            self.assertEqual(tier.pk, snap['tier_id'])

    def test_excluded_and_incomplete_have_no_wholesale(self):
        self._run()
        for article in EXCLUDED_ARTICLES:
            product = Product.objects.get(seller_profile=self.seller, article=article)
            self.assertFalse(product.publish_to_sellers)
            self.assertFalse(product.price_tiers.exists())
            self.assertFalse(has_public_wholesale_offer(product, self.seller))
            self.assertEqual(product.price, 1800)

    def test_legacy_duplicates_cleaned(self):
        self._run()
        for spec in LEGACY_NULL_DUPLICATES:
            canonical = Product.objects.get(
                seller_profile=self.seller,
                article=spec['article'],
            )
            self.assertEqual(canonical.status, 'active')
            self.assertTrue(canonical.publish_to_sellers)
        self.assertFalse(Product.objects.filter(pk=self.legacy_cr.pk).exists())
        self.assertFalse(Product.objects.filter(pk=self.legacy_txl.pk).exists())
        self.assertFalse(Product.objects.filter(pk=self.legacy_cs35.pk).exists())
        self.assertEqual(
            Product.objects.filter(article='1109190CR01').count(),
            1,
        )
        self.assertEqual(
            Product.objects.filter(article='151000025AA').count(),
            1,
        )
        self.assertEqual(
            Product.objects.filter(article='S1010140400').count(),
            1,
        )

    def test_legacy_duplicate_with_order_is_hidden_not_deleted(self):
        order = Order.objects.create(
            customer_name='Buyer',
            customer_phone='77000000001',
            total_price=2750,
            delivery_method=Order.DELIVERY_PICKUP,
        )
        OrderItem.objects.create(
            order=order,
            product=self.legacy_cr,
            quantity=1,
            price_at_purchase=2750,
        )
        self._run()
        self.legacy_cr.refresh_from_db()
        self.assertEqual(self.legacy_cr.status, 'hidden')
        self.assertFalse(self.legacy_cr.publish_to_sellers)
        self.assertFalse(self.legacy_cr.price_tiers.exists())
        self.assertTrue(
            Product.objects.filter(
                seller_profile=self.seller,
                article='1109190CR01',
                status='active',
            ).exists()
        )

    def test_xlsx_has_47_product_rows(self):
        self._run()
        response = self.client.get(
            reverse('public_seller_wholesale_price', kwargs={'slug': self.seller.slug})
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        articles = [
            row[0]
            for row in workbook['Прайс'].iter_rows(min_row=2, values_only=True)
            if row[0]
        ]
        self.assertEqual(len(articles), 47)
        self.assertEqual(
            set(articles),
            set(ORIGINAL_WHOLESALE_ARTICLES) | set(STAGE1_AIR_FILTERS),
        )

    def test_terms_and_whatsapp_untouched(self):
        self._run()
        self.seller.refresh_from_db()
        self.assertEqual(self.seller.wholesale_min_order_qty, 10)
        self.terms.refresh_from_db()
        self.assertEqual(self.terms.vat_mode, self.terms_snapshot['vat_mode'])
        self.assertEqual(
            self.terms.prepayment_percent,
            self.terms_snapshot['prepayment_percent'],
        )
        if self.template_snapshot is not None:
            template = MarketingWhatsAppTemplate.objects.get(
                meta_template_name='zpt_ag_parts_wholesale_v1',
                language_code='ru',
            )
            self.assertEqual(template.meta_status, self.template_snapshot['meta_status'])
            self.assertEqual(template.meta_template_id, self.template_snapshot['meta_template_id'])
            self.assertEqual(template.body_text, self.template_snapshot['body'])
            self.assertEqual(template.header_text, self.template_snapshot['header'])

    def test_missing_seller_is_noop(self):
        self.seller.slug = 'ag-parts-other'
        self.seller.save(update_fields=['slug'])
        STAGE1.upsert_ag_parts_stage1_air_filters(apps, None)
        self.assertFalse(
            Product.objects.filter(article='J691109111').exists()
        )
