from importlib import import_module

from django.apps import apps
from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from io import BytesIO
from openpyxl import load_workbook

from catalog.ag_parts_air_filters import (
    APPROVED_AIR_FILTER_ARTICLES,
    EXCLUDED_ARTICLES,
    ORIGINAL_WHOLESALE_ARTICLES,
    STAGE1_AIR_FILTERS,
    STAGE2_AIR_FILTERS,
)
from catalog.models import Product, ProductPriceTier, SellerProfile, SellerWholesaleTerms
from catalog.product_photo_import import (
    plan_product_photo_import,
    seller_photo_whitelist,
)
from catalog.wholesale import has_public_wholesale_offer, wholesale_products_qs
from marketing.models import MarketingWhatsAppTemplate
from PIL import Image

import zipfile


STAGE3 = import_module('catalog.migrations.0029_ag_parts_stage3_air_filter')

UNRESOLVED = ('PBC1109610', '1109110XP6EXA')
ARTICLE = '1109110XKV08A'


def _make_seller(username, name, phone, **kwargs):
    user = User.objects.create_user(username=username, password='secret12345')
    defaults = {
        'user': user,
        'name': name,
        'phone': phone,
        'city': 'Алматы',
        'address': 'г. Алматы, ул. Тестовая, 1',
        'wholesale_enabled': True,
        'wholesale_min_order_qty': 10,
    }
    defaults.update(kwargs)
    return SellerProfile.objects.create(**defaults)


class AgPartsStage3AirFilterTests(TestCase):
    def setUp(self):
        self.seller = _make_seller('ag-parts-air3', 'AG Parts', '77771360740')
        self.assertEqual(self.seller.slug, 'ag-parts')
        self.terms = SellerWholesaleTerms.objects.create(
            seller=self.seller,
            vat_mode=SellerWholesaleTerms.VAT_INCLUDED,
            prepayment_percent=100,
            pickup_city='Алматы',
            primary_carrier='DPD Kazakhstan',
        )
        self.previous = []
        for index, article in enumerate(sorted(ORIGINAL_WHOLESALE_ARTICLES)):
            product = self._owned(
                title=f'Approved {article}',
                article=article,
                slug=f'ag-orig-{index}',
                price=2000 + index,
                stock_qty=11 + index,
                publish_to_sellers=True,
                publish_to_kaspi=(index == 0),
            )
            ProductPriceTier.objects.create(
                product=product, min_qty=1, price=400 + index
            )
            self.previous.append(product)
        for index, (article, (retail, wholesale)) in enumerate(
            sorted(STAGE1_AIR_FILTERS.items())
        ):
            product = self._owned(
                title=f'Воздушный фильтр {article}',
                article=article,
                slug=f'ag-air1-{index}',
                price=retail,
                publish_to_sellers=True,
            )
            ProductPriceTier.objects.create(
                product=product, min_qty=1, price=wholesale
            )
            self.previous.append(product)
        for index, (article, spec) in enumerate(sorted(STAGE2_AIR_FILTERS.items())):
            product = self._owned(
                title=spec['title'],
                article=article,
                slug=f'ag-air2-{index}',
                price=spec['retail'],
                publish_to_sellers=True,
            )
            ProductPriceTier.objects.create(
                product=product, min_qty=1, price=spec['wholesale']
            )
            self.previous.append(product)

        self.legacy = Product.objects.create(
            pk=2046,
            title='Воздушный фильтр',
            article=ARTICLE,
            slug='ag-legacy-xkv08a',
            price=2310,
            seller_name='AG Parts shop',
            seller_profile=None,
            whatsapp_number='+77771360740',
            status='active',
            city='Алматы',
            publish_to_sellers=False,
            publish_to_kaspi=True,
            stock_qty=None,
        )
        for article in UNRESOLVED:
            self._owned(
                title=f'Unresolved {article}',
                article=article,
                slug=f'ag-unres-{article.lower()}',
                price=2310,
                stock_qty=15 if article == 'PBC1109610' else None,
                publish_to_sellers=False,
            )
        self.snapshot = self._snapshot_previous()
        self.terms_snapshot = {
            'vat_mode': self.terms.vat_mode,
            'prepayment_percent': self.terms.prepayment_percent,
            'pickup_city': self.terms.pickup_city,
            'primary_carrier': self.terms.primary_carrier,
        }
        self.template_snapshot = self._template_snapshot()

    def _owned(self, **kwargs):
        defaults = {
            'seller_name': self.seller.name,
            'seller_profile': self.seller,
            'whatsapp_number': '+77771360740',
            'status': 'active',
            'city': 'Алматы',
            'publish_to_sellers': False,
            'publish_to_kaspi': False,
        }
        defaults.update(kwargs)
        return Product.objects.create(**defaults)

    def _snapshot_previous(self):
        data = {}
        for product in self.previous:
            product.refresh_from_db()
            tier = product.price_tiers.get(min_qty=1, is_active=True)
            data[product.article] = {
                'price': product.price,
                'stock_qty': product.stock_qty,
                'publish_to_sellers': product.publish_to_sellers,
                'publish_to_kaspi': product.publish_to_kaspi,
                'status': product.status,
                'title': product.title,
                'tier_price': tier.price,
                'tier_id': tier.pk,
            }
        return data

    def _template_snapshot(self):
        template = MarketingWhatsAppTemplate.objects.filter(
            meta_template_name='zpt_ag_parts_wholesale_v1',
            language_code='ru',
        ).first()
        if template is None:
            return None
        return {
            'meta_status': template.meta_status,
            'meta_template_id': template.meta_template_id,
            'body': template.body_text,
            'header': template.header_text,
        }

    def _run(self):
        STAGE3.upsert_ag_parts_stage3_air_filter(apps, None)
        STAGE3.upsert_ag_parts_stage3_air_filter(apps, None)

    def test_reuses_legacy_product_without_duplicate(self):
        self._run()
        rows = list(Product.objects.filter(article=ARTICLE).order_by('id'))
        self.assertEqual(len(rows), 1)
        product = rows[0]
        self.assertEqual(product.pk, 2046)
        self.assertEqual(product.pk, self.legacy.pk)
        self.assertEqual(product.seller_profile_id, self.seller.pk)
        self.assertEqual(product.seller_name, self.seller.name)
        self.assertEqual(product.status, 'active')
        self.assertTrue(product.publish_to_sellers)
        self.assertEqual(product.price, 2310)
        self.assertTrue(product.publish_to_kaspi)
        self.assertIsNone(product.stock_qty)
        self.assertEqual(product.title, 'Воздушный фильтр 1109110XKV08A')
        tiers = list(product.price_tiers.filter(is_active=True))
        self.assertEqual(len(tiers), 1)
        self.assertEqual(tiers[0].min_qty, 1)
        self.assertEqual(tiers[0].price, 490)
        self.assertTrue(has_public_wholesale_offer(product, self.seller))

    def test_total_wholesale_51_air_25(self):
        self._run()
        wholesale = list(
            wholesale_products_qs(self.seller).order_by('article').values_list(
                'article', flat=True
            )
        )
        self.assertEqual(len(wholesale), 51)
        self.assertEqual(len(self.snapshot), 50)
        air = [article for article in wholesale if article in APPROVED_AIR_FILTER_ARTICLES]
        self.assertEqual(len(air), 25)
        self.assertEqual(set(air), set(APPROVED_AIR_FILTER_ARTICLES))
        self.assertIn(ARTICLE, wholesale)

    def test_previous_50_and_unresolved_untouched(self):
        self._run()
        for product in self.previous:
            product.refresh_from_db()
            snap = self.snapshot[product.article]
            self.assertEqual(product.price, snap['price'])
            self.assertEqual(product.stock_qty, snap['stock_qty'])
            self.assertEqual(product.publish_to_sellers, snap['publish_to_sellers'])
            self.assertEqual(product.publish_to_kaspi, snap['publish_to_kaspi'])
            self.assertEqual(product.status, snap['status'])
            self.assertEqual(product.title, snap['title'])
            tier = product.price_tiers.get(min_qty=1, is_active=True)
            self.assertEqual(tier.price, snap['tier_price'])
            self.assertEqual(tier.pk, snap['tier_id'])
        self.assertEqual(set(EXCLUDED_ARTICLES), set(UNRESOLVED))
        for article in UNRESOLVED:
            product = Product.objects.get(seller_profile=self.seller, article=article)
            self.assertFalse(product.publish_to_sellers)
            self.assertFalse(product.price_tiers.exists())
            self.assertFalse(has_public_wholesale_offer(product, self.seller))

    def test_photo_alias_kvo8a(self):
        self._run()
        jpeg = BytesIO()
        Image.new('RGB', (8, 8), (10, 20, 30)).save(jpeg, format='JPEG')
        buffer = BytesIO()
        with zipfile.ZipFile(buffer, 'w') as bundle:
            bundle.writestr('1109110XKVO8A/front.jpg', jpeg.getvalue())
            bundle.writestr('PBC1109610/skip.jpg', jpeg.getvalue())
        rows = {
            row.folder_name: row
            for row in plan_product_photo_import(self.seller, buffer.getvalue())
        }
        self.assertEqual(rows['1109110XKVO8A'].article, ARTICLE)
        self.assertEqual(rows['1109110XKVO8A'].alias_used, '1109110XKVO8A')
        self.assertEqual(rows['1109110XKVO8A'].display_status, 'matched')
        self.assertEqual(rows['1109110XKVO8A'].product.pk, 2046)
        self.assertEqual(rows['PBC1109610'].display_status, 'skipped')
        self.assertEqual(len(APPROVED_AIR_FILTER_ARTICLES), 25)
        self.assertEqual(seller_photo_whitelist(self.seller), APPROVED_AIR_FILTER_ARTICLES)
        self.assertNotIn('PBC1109610', APPROVED_AIR_FILTER_ARTICLES)
        self.assertNotIn('1109110XP6EXA', APPROVED_AIR_FILTER_ARTICLES)

    def test_xlsx_has_51_rows(self):
        self._run()
        response = self.client.get(
            reverse('public_seller_wholesale_price', kwargs={'slug': self.seller.slug})
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        articles = [
            row[0]
            for row in workbook['Прайс'].iter_rows(min_row=2, values_only=True)
            if row[0]
        ]
        self.assertEqual(len(articles), 51)
        rec = next(row for row in workbook['Прайс'].iter_rows(min_row=2, values_only=True) if row[0] == ARTICLE)
        self.assertEqual(rec[4], 2310)
        self.assertEqual(rec[5], 490)

    def test_terms_and_whatsapp_untouched(self):
        self._run()
        self.seller.refresh_from_db()
        self.assertEqual(self.seller.wholesale_min_order_qty, 10)
        self.terms.refresh_from_db()
        self.assertEqual(self.terms.vat_mode, self.terms_snapshot['vat_mode'])
        if self.template_snapshot is not None:
            template = MarketingWhatsAppTemplate.objects.get(
                meta_template_name='zpt_ag_parts_wholesale_v1',
                language_code='ru',
            )
            self.assertEqual(template.meta_status, self.template_snapshot['meta_status'])
            self.assertEqual(template.body_text, self.template_snapshot['body'])

    def test_prefers_production_pk_among_legacy_rows(self):
        other = Product.objects.create(
            title='Extra unbound',
            article=ARTICLE,
            slug='ag-legacy-xkv08a-extra',
            price=3149,
            seller_name='AG Parts',
            seller_profile=None,
            whatsapp_number='+77771360740',
            status='hidden',
            city='Алматы',
            publish_to_sellers=False,
        )
        self.assertNotEqual(other.pk, 2046)
        self._run()
        product = Product.objects.get(pk=2046)
        self.assertEqual(product.seller_profile_id, self.seller.pk)
        self.assertEqual(product.price, 2310)
        self.assertEqual(product.price_tiers.get(min_qty=1, is_active=True).price, 490)
        other.refresh_from_db()
        self.assertIsNone(other.seller_profile_id)
        self.assertFalse(other.publish_to_sellers)
        self.assertFalse(other.price_tiers.exists())
        self.assertEqual(other.price, 3149)
        self.assertEqual(Product.objects.filter(article=ARTICLE).count(), 2)

    def test_creates_when_missing(self):
        Product.objects.filter(article=ARTICLE).delete()
        STAGE3.upsert_ag_parts_stage3_air_filter(apps, None)
        product = Product.objects.get(article=ARTICLE, seller_profile=self.seller)
        self.assertNotEqual(product.pk, 2046)
        self.assertEqual(product.price, 2310)
        self.assertEqual(product.price_tiers.get(min_qty=1).price, 490)
        self.assertFalse(product.publish_to_kaspi)
        self.assertIsNone(product.stock_qty)
        self.assertEqual(Product.objects.filter(article=ARTICLE).count(), 1)
