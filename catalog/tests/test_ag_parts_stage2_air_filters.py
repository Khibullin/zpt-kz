from importlib import import_module

from django.apps import apps
from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from io import BytesIO
from openpyxl import load_workbook

from catalog.ag_parts_air_filters import (
    APPROVED_AIR_FILTER_ARTICLES,
    EXCLUDED_ARTICLES,
    ORIGINAL_WHOLESALE_ARTICLES,
    STAGE1_AIR_FILTERS,
    STAGE2_AIR_FILTERS,
)
from catalog.applicability import serialize_plain_list
from catalog.models import Product, ProductPriceTier, SellerProfile, SellerWholesaleTerms
from catalog.product_photo_import import seller_photo_whitelist
from catalog.wholesale import has_public_wholesale_offer, wholesale_products_qs
from marketing.models import MarketingWhatsAppTemplate


STAGE2 = import_module('catalog.migrations.0028_ag_parts_stage2_air_filters')

UNRESOLVED = ('1109110XKV08A', 'PBC1109610', '1109110XP6EXA')


def _make_seller(username, name, phone, **kwargs):
    user = User.objects.create_user(username=username, password='secret12345')
    defaults = {
        'user': user,
        'name': name,
        'phone': phone,
        'city': 'Алматы',
        'address': 'г. Алматы, ул. Тестовая, 1',
        'wholesale_enabled': True,
        'wholesale_min_order_qty': 10,
    }
    defaults.update(kwargs)
    return SellerProfile.objects.create(**defaults)


class AgPartsStage2AirFiltersTests(TestCase):
    def setUp(self):
        self.seller = _make_seller('ag-parts-air2', 'AG Parts', '77771360740')
        self.assertEqual(self.seller.slug, 'ag-parts')
        self.terms = SellerWholesaleTerms.objects.create(
            seller=self.seller,
            vat_mode=SellerWholesaleTerms.VAT_INCLUDED,
            prepayment_percent=100,
            pickup_city='Алматы',
            primary_carrier='DPD Kazakhstan',
        )
        self.previous = []
        for index, article in enumerate(sorted(ORIGINAL_WHOLESALE_ARTICLES)):
            product = self._product(
                title=f'Approved {article}',
                article=article,
                slug=f'ag-orig-{index}',
                price=2000 + index,
                stock_qty=11 + index,
                publish_to_sellers=True,
                publish_to_kaspi=(index == 0),
            )
            ProductPriceTier.objects.create(
                product=product, min_qty=1, price=400 + index
            )
            self.previous.append(product)
        for index, (article, (retail, wholesale)) in enumerate(
            sorted(STAGE1_AIR_FILTERS.items())
        ):
            product = self._product(
                title=f'Воздушный фильтр {article}',
                article=article,
                slug=f'ag-air1-{index}',
                price=retail,
                stock_qty=5 if article == 'T151109111' else None,
                publish_to_sellers=True,
                publish_to_kaspi=(article == 'T151109111'),
            )
            ProductPriceTier.objects.create(
                product=product, min_qty=1, price=wholesale
            )
            self.previous.append(product)

        self.existing_oem = self._product(
            title='Old air 1064000180',
            article='1064000180',
            slug='ag-air2-1064',
            price=1150,
            stock_qty=None,
            publish_to_sellers=False,
            publish_to_kaspi=False,
        )
        self.existing_jac = self._product(
            title='Old JAC S5 air',
            article='1109130U1510',
            slug='ag-air2-jac-s5',
            price=1600,
            stock_qty=7,
            publish_to_sellers=False,
            publish_to_kaspi=True,
        )
        for article in UNRESOLVED:
            self._product(
                title=f'Unresolved {article}',
                article=article,
                slug=f'ag-unres-{article.lower()}',
                price=2310,
                stock_qty=15 if article == 'PBC1109610' else None,
                publish_to_sellers=False,
            )

        self.snapshot = self._snapshot_previous()
        self.terms_snapshot = {
            'vat_mode': self.terms.vat_mode,
            'prepayment_percent': self.terms.prepayment_percent,
            'pickup_city': self.terms.pickup_city,
            'primary_carrier': self.terms.primary_carrier,
        }
        self.template_snapshot = self._template_snapshot()

    def _product(self, **kwargs):
        defaults = {
            'seller_name': self.seller.name,
            'seller_profile': self.seller,
            'whatsapp_number': '+77771360740',
            'status': 'active',
            'city': 'Алматы',
            'publish_to_sellers': False,
            'publish_to_kaspi': False,
        }
        defaults.update(kwargs)
        return Product.objects.create(**defaults)

    def _snapshot_previous(self):
        data = {}
        for product in self.previous:
            product.refresh_from_db()
            tier = product.price_tiers.get(min_qty=1, is_active=True)
            data[product.article] = {
                'price': product.price,
                'stock_qty': product.stock_qty,
                'publish_to_sellers': product.publish_to_sellers,
                'publish_to_kaspi': product.publish_to_kaspi,
                'status': product.status,
                'title': product.title,
                'tier_price': tier.price,
                'tier_id': tier.pk,
            }
        return data

    def _template_snapshot(self):
        template = MarketingWhatsAppTemplate.objects.filter(
            meta_template_name='zpt_ag_parts_wholesale_v1',
            language_code='ru',
        ).first()
        if template is None:
            return None
        return {
            'meta_status': template.meta_status,
            'meta_template_id': template.meta_template_id,
            'body': template.body_text,
            'header': template.header_text,
        }

    def _run(self):
        STAGE2.upsert_ag_parts_stage2_air_filters(apps, None)
        STAGE2.upsert_ag_parts_stage2_air_filters(apps, None)

    def test_three_resolved_sku_and_total_wholesale(self):
        self._run()
        wholesale = list(
            wholesale_products_qs(self.seller).order_by('article').values_list(
                'article', flat=True
            )
        )
        self.assertEqual(len(wholesale), 50)
        self.assertEqual(
            set(wholesale),
            set(ORIGINAL_WHOLESALE_ARTICLES)
            | set(STAGE1_AIR_FILTERS)
            | set(STAGE2_AIR_FILTERS),
        )
        air = [
            article for article in wholesale
            if article in APPROVED_AIR_FILTER_ARTICLES
        ]
        self.assertEqual(len(air), 24)
        self.assertEqual(set(air), set(APPROVED_AIR_FILTER_ARTICLES))
        self.assertEqual(set(STAGE2_AIR_FILTERS), {'1064000180', '1109130U1510', 'F081109111HD'})

    def test_exact_prices_flags_titles_and_stock(self):
        self._run()
        for article, spec in STAGE2_AIR_FILTERS.items():
            product = Product.objects.get(seller_profile=self.seller, article=article)
            self.assertEqual(product.status, 'active')
            self.assertEqual(product.seller_profile_id, self.seller.pk)
            self.assertTrue(product.publish_to_sellers)
            self.assertEqual(product.price, spec['retail'])
            tiers = list(product.price_tiers.filter(is_active=True))
            self.assertEqual(len(tiers), 1)
            self.assertEqual(tiers[0].min_qty, 1)
            self.assertEqual(tiers[0].price, spec['wholesale'])
            self.assertTrue(has_public_wholesale_offer(product, self.seller))
            self.assertEqual(product.title, spec['title'])
            if spec['compatibility']:
                self.assertEqual(product.compatibility, spec['compatibility'])
            if spec['oem_cross_references']:
                self.assertEqual(
                    product.oem_cross_references,
                    serialize_plain_list(spec['oem_cross_references']),
                )
        self.existing_oem.refresh_from_db()
        self.existing_jac.refresh_from_db()
        self.assertIsNone(self.existing_oem.stock_qty)
        self.assertFalse(self.existing_oem.publish_to_kaspi)
        self.assertEqual(self.existing_jac.stock_qty, 7)
        self.assertTrue(self.existing_jac.publish_to_kaspi)
        created = Product.objects.get(seller_profile=self.seller, article='F081109111HD')
        self.assertIsNone(created.stock_qty)
        self.assertFalse(created.publish_to_kaspi)

    def test_previous_47_unchanged(self):
        self._run()
        self.assertEqual(len(self.snapshot), 47)
        for product in self.previous:
            product.refresh_from_db()
            snap = self.snapshot[product.article]
            self.assertEqual(product.price, snap['price'])
            self.assertEqual(product.stock_qty, snap['stock_qty'])
            self.assertEqual(product.publish_to_sellers, snap['publish_to_sellers'])
            self.assertEqual(product.publish_to_kaspi, snap['publish_to_kaspi'])
            self.assertEqual(product.status, snap['status'])
            self.assertEqual(product.title, snap['title'])
            tier = product.price_tiers.get(min_qty=1, is_active=True)
            self.assertEqual(tier.price, snap['tier_price'])
            self.assertEqual(tier.pk, snap['tier_id'])

    def test_unresolved_still_without_wholesale(self):
        self._run()
        self.assertEqual(set(EXCLUDED_ARTICLES), set(UNRESOLVED))
        for article in UNRESOLVED:
            product = Product.objects.get(seller_profile=self.seller, article=article)
            self.assertFalse(product.publish_to_sellers)
            self.assertFalse(product.price_tiers.exists())
            self.assertFalse(has_public_wholesale_offer(product, self.seller))
            self.assertEqual(product.price, 2310)
        pbc = Product.objects.get(seller_profile=self.seller, article='PBC1109610')
        self.assertEqual(pbc.stock_qty, 15)

    def test_photo_whitelist_is_24(self):
        self.assertEqual(len(APPROVED_AIR_FILTER_ARTICLES), 24)
        self.assertEqual(seller_photo_whitelist(self.seller), APPROVED_AIR_FILTER_ARTICLES)
        for article in STAGE2_AIR_FILTERS:
            self.assertIn(article, APPROVED_AIR_FILTER_ARTICLES)
        for article in UNRESOLVED:
            self.assertNotIn(article, APPROVED_AIR_FILTER_ARTICLES)

    def test_xlsx_has_50_product_rows(self):
        self._run()
        response = self.client.get(
            reverse('public_seller_wholesale_price', kwargs={'slug': self.seller.slug})
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        articles = [
            row[0]
            for row in workbook['Прайс'].iter_rows(min_row=2, values_only=True)
            if row[0]
        ]
        self.assertEqual(len(articles), 50)
        self.assertEqual(
            set(articles),
            set(ORIGINAL_WHOLESALE_ARTICLES)
            | set(STAGE1_AIR_FILTERS)
            | set(STAGE2_AIR_FILTERS),
        )

    def test_terms_and_whatsapp_untouched(self):
        self._run()
        self.seller.refresh_from_db()
        self.assertEqual(self.seller.wholesale_min_order_qty, 10)
        self.terms.refresh_from_db()
        self.assertEqual(self.terms.vat_mode, self.terms_snapshot['vat_mode'])
        self.assertEqual(
            self.terms.prepayment_percent,
            self.terms_snapshot['prepayment_percent'],
        )
        if self.template_snapshot is not None:
            template = MarketingWhatsAppTemplate.objects.get(
                meta_template_name='zpt_ag_parts_wholesale_v1',
                language_code='ru',
            )
            self.assertEqual(template.meta_status, self.template_snapshot['meta_status'])
            self.assertEqual(template.meta_template_id, self.template_snapshot['meta_template_id'])
            self.assertEqual(template.body_text, self.template_snapshot['body'])
            self.assertEqual(template.header_text, self.template_snapshot['header'])

    def test_missing_seller_is_noop(self):
        self.seller.slug = 'ag-parts-other'
        self.seller.save(update_fields=['slug'])
        STAGE2.upsert_ag_parts_stage2_air_filters(apps, None)
        self.assertFalse(
            Product.objects.filter(article='F081109111HD').exists()
        )
        self.existing_jac.refresh_from_db()
        self.assertFalse(self.existing_jac.publish_to_sellers)
        self.assertFalse(self.existing_jac.price_tiers.exists())
