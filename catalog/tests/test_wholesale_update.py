from io import BytesIO
from unittest.mock import patch

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import TYPE_FORMULA

from catalog.models import (
    CatalogImportBatch,
    CatalogImportItem,
    Product,
    ProductFulfillment,
    ProductPriceTier,
    SellerProfile,
)
from catalog.wholesale_export import XLSX_CONTENT_TYPE
from catalog.wholesale_update import (
    DOWNLOAD_HEADERS,
    MAX_UPLOAD_BYTES,
    WHOLESALE_GTE_RETAIL_WARNING,
    plan_wholesale_update,
    sha256_bytes,
    wholesale_update_filename,
)


RETAIL = 2189
WHOLESALE = 310
COST = 87654321


def _make_seller(username, name='Update Shop', **kwargs):
    user = User.objects.create_user(username=username, password='secret12345')
    defaults = {
        'user': user,
        'name': name,
        'phone': '77770001122',
        'city': 'Алматы',
        'address': 'Тестовая, 1',
        'wholesale_enabled': True,
        'wholesale_min_order_qty': 10,
    }
    defaults.update(kwargs)
    return SellerProfile.objects.create(**defaults)


def _xlsx_bytes(rows, headers=None):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers or DOWNLOAD_HEADERS)
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class WholesaleUpdateAdminTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(
            'wh-admin', 'admin@test.local', 'secret12345'
        )
        self.plain = User.objects.create_user('buyer', password='secret12345')
        self.seller = _make_seller('wh-update-owner', 'AG Parts')
        self.other = _make_seller('wh-update-other', 'Other Shop', phone='77000000001')
        self.product = Product.objects.create(
            title='Салонный фильтр JAC',
            price=RETAIL,
            cost_price=COST,
            seller_name=self.seller.name,
            seller_profile=self.seller,
            whatsapp_number='+77770001122',
            status='active',
            publish_to_sellers=True,
            city='Алматы',
            article='8114010U8520',
            slug='wh-upd-jac',
        )
        ProductPriceTier.objects.create(
            product=self.product, min_qty=1, price=WHOLESALE
        )
        self.other_product = Product.objects.create(
            title='Чужой товар',
            price=1111,
            seller_name=self.other.name,
            seller_profile=self.other,
            whatsapp_number='+77000000001',
            status='active',
            publish_to_sellers=True,
            city='Алматы',
            article='OTHER-ART',
            slug='wh-upd-other',
        )
        ProductPriceTier.objects.create(
            product=self.other_product, min_qty=1, price=500
        )

    def _download_url(self, seller=None):
        seller = seller or self.seller
        return reverse(
            'admin:catalog_sellerprofile_wholesale_download',
            args=[seller.pk],
        )

    def _upload_url(self, seller=None):
        seller = seller or self.seller
        return reverse(
            'admin:catalog_sellerprofile_wholesale_update',
            args=[seller.pk],
        )

    def _apply_url(self, batch, seller=None):
        seller = seller or self.seller
        return reverse(
            'admin:catalog_sellerprofile_wholesale_apply',
            args=[seller.pk, batch.pk],
        )

    def _login_admin(self):
        self.client.force_login(self.admin)

    def _upload(self, payload, name='update.xlsx'):
        self._login_admin()
        uploaded = SimpleUploadedFile(
            name,
            payload,
            content_type=XLSX_CONTENT_TYPE,
        )
        return self.client.post(self._upload_url(), {'file': uploaded})

    def test_staff_can_download_current_file(self):
        self.product.stock_qty = 27
        self.product.save(update_fields=['stock_qty'])
        self._login_admin()
        response = self.client.get(self._download_url())
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], XLSX_CONTENT_TYPE)
        filename = wholesale_update_filename(self.seller)
        self.assertIn(filename, response['Content-Disposition'])
        workbook = load_workbook(BytesIO(response.content))
        headers = [cell.value for cell in next(workbook.active.iter_rows(max_row=1))]
        self.assertEqual(headers, DOWNLOAD_HEADERS)
        row = next(workbook.active.iter_rows(min_row=2, values_only=True))
        self.assertEqual(row[0], self.product.article)
        self.assertEqual(row[2], RETAIL)
        self.assertEqual(row[3], WHOLESALE)
        self.assertEqual(row[4], 27)
        blob = ' '.join(str(value) for value in headers + list(row))
        self.assertNotIn(str(COST), blob)
        self.assertNotIn('cost_price', blob)

    def test_anonymous_and_nonstaff_cannot_download(self):
        self.assertEqual(self.client.get(self._download_url()).status_code, 302)
        self.client.force_login(self.plain)
        self.assertEqual(self.client.get(self._download_url()).status_code, 302)

    def test_empty_stock_cell_in_download(self):
        self._login_admin()
        workbook = load_workbook(BytesIO(self.client.get(self._download_url()).content))
        row = next(workbook.active.iter_rows(min_row=2, values_only=True))
        self.assertIn(row[4], (None, ''))

    def test_preview_does_not_mutate_product(self):
        payload = _xlsx_bytes([[self.product.article, 'NEW TITLE', 2200, 320, 27]])
        response = self._upload(payload)
        self.assertEqual(response.status_code, 302)
        self.product.refresh_from_db()
        self.assertEqual(self.product.title, 'Салонный фильтр JAC')
        self.assertEqual(self.product.price, RETAIL)
        self.assertIsNone(self.product.stock_qty)
        self.assertEqual(
            self.product.price_tiers.get(min_qty=1).price, WHOLESALE
        )
        batch = CatalogImportBatch.objects.get()
        self.assertEqual(batch.source, CatalogImportBatch.SOURCE_WHOLESALE_UPDATE)
        self.assertEqual(batch.mode, CatalogImportBatch.MODE_DRY_RUN)
        self.assertEqual(batch.source_scope, CatalogImportBatch.SCOPE_PARTIAL)
        self.assertEqual(batch.uploaded_by, self.admin)
        html = self.client.get(response.url).content.decode('utf-8')
        self.assertIn('изменится', html)
        self.assertIn('Применить изменения', html)

    def test_retail_only_wholesale_only_stock_only_and_blank(self):
        rows = plan_wholesale_update(
            self.seller,
            _xlsx_bytes([[self.product.article, 'ignored', 2200, '', '']]),
        )
        self.assertEqual(rows[0].action, CatalogImportItem.ACTION_UPDATED)
        self.assertIn('price', rows[0].changed_fields)
        self.assertNotIn('wholesale_price', rows[0].changed_fields)
        self.assertNotIn('stock_qty', rows[0].changed_fields)

        rows = plan_wholesale_update(
            self.seller,
            _xlsx_bytes([[self.product.article, '', '', 320, '']]),
        )
        self.assertIn('wholesale_price', rows[0].changed_fields)
        self.assertNotIn('price', rows[0].changed_fields)

        rows = plan_wholesale_update(
            self.seller,
            _xlsx_bytes([[self.product.article, '', '', '', 0]]),
        )
        self.assertEqual(rows[0].changed_fields['stock_qty']['new'], 0)
        self.assertIsNone(rows[0].changed_fields['stock_qty']['old'])

        rows = plan_wholesale_update(
            self.seller,
            _xlsx_bytes([[self.product.article, 'x', '', '', '']]),
        )
        self.assertEqual(rows[0].action, CatalogImportItem.ACTION_UNCHANGED)

    def test_mixed_row_and_unknown_duplicate_wrong_seller(self):
        rows = plan_wholesale_update(
            self.seller,
            _xlsx_bytes([[self.product.article, '', 2200, 320, 8]]),
        )
        self.assertEqual(rows[0].action, CatalogImportItem.ACTION_UPDATED)

        rows = plan_wholesale_update(
            self.seller,
            _xlsx_bytes([['NO-SUCH', '', 1, 1, 1]]),
        )
        self.assertEqual(rows[0].action, CatalogImportItem.ACTION_CONFLICT)

        rows = plan_wholesale_update(
            self.seller,
            _xlsx_bytes([
                [self.product.article, '', 2200, '', ''],
                [self.product.article.lower(), '', 2300, '', ''],
            ]),
        )
        self.assertTrue(
            all(row.action == CatalogImportItem.ACTION_CONFLICT for row in rows)
        )

        rows = plan_wholesale_update(
            self.seller,
            _xlsx_bytes([[self.other_product.article, '', 2000, 100, 1]]),
        )
        self.assertEqual(rows[0].action, CatalogImportItem.ACTION_CONFLICT)
        self.assertIsNone(rows[0].product)

    def test_fractional_price_and_formula_and_limits(self):
        rows = plan_wholesale_update(
            self.seller,
            _xlsx_bytes([[self.product.article, '', 2200.5, '', '']]),
        )
        self.assertEqual(rows[0].action, CatalogImportItem.ACTION_ERROR)
        self.assertTrue(rows[0].errors)

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(DOWNLOAD_HEADERS)
        sheet.append([self.product.article, '', 2200, 320, 1])
        sheet['C2'] = '=1+1'
        self.assertEqual(sheet['C2'].data_type, TYPE_FORMULA)
        buffer = BytesIO()
        workbook.save(buffer)
        rows = plan_wholesale_update(self.seller, buffer.getvalue())
        self.assertEqual(rows[0].action, CatalogImportItem.ACTION_ERROR)
        self.assertTrue(any('формул' in err.lower() for err in rows[0].errors))

        with patch('catalog.wholesale_update.MAX_DATA_ROWS', 1):
            from catalog.wholesale_update import WholesaleUpdateError
            with self.assertRaises(WholesaleUpdateError):
                plan_wholesale_update(
                    self.seller,
                    _xlsx_bytes([
                        [self.product.article, '', 2200, '', ''],
                        ['X', '', 1, '', ''],
                    ]),
                )

        self._login_admin()
        huge = SimpleUploadedFile(
            'huge.xlsx',
            b'PK\x03\x04' + b'0' * (MAX_UPLOAD_BYTES + 10),
            content_type=XLSX_CONTENT_TYPE,
        )
        response = self.client.post(self._upload_url(), {'file': huge})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '5 МБ')

    def test_wholesale_gte_retail_warning_not_blocking(self):
        rows = plan_wholesale_update(
            self.seller,
            _xlsx_bytes([[self.product.article, '', 300, 300, '']]),
        )
        self.assertEqual(rows[0].action, CatalogImportItem.ACTION_UPDATED)
        self.assertIn(WHOLESALE_GTE_RETAIL_WARNING, rows[0].warnings)

    def test_apply_updates_allowed_fields_atomically(self):
        payload = _xlsx_bytes([[self.product.article, 'IGNORE TITLE', 2200, 320, 27]])
        self._upload(payload)
        batch = CatalogImportBatch.objects.get(mode=CatalogImportBatch.MODE_DRY_RUN)
        apply = self.client.post(self._apply_url(batch), {'confirm': '1'})
        self.assertEqual(apply.status_code, 200)
        self.assertContains(apply, 'Изменения применены')
        self.product.refresh_from_db()
        self.assertEqual(self.product.title, 'Салонный фильтр JAC')
        self.assertEqual(self.product.price, 2200)
        self.assertEqual(self.product.stock_qty, 27)
        self.assertEqual(self.product.cost_price, COST)
        self.assertTrue(self.product.publish_to_sellers)
        self.assertEqual(self.product.status, 'active')
        self.assertEqual(self.product.price_tiers.get(min_qty=1).price, 320)
        write = CatalogImportBatch.objects.get(mode=CatalogImportBatch.MODE_WRITE)
        self.assertEqual(write.source, CatalogImportBatch.SOURCE_WHOLESALE_UPDATE)
        self.assertEqual(write.status, CatalogImportBatch.STATUS_SUCCESS)
        self.assertEqual(write.applied_by, self.admin)
        item = write.items.get()
        self.assertEqual(item.changed_fields['price'], {'old': RETAIL, 'new': 2200})
        self.assertEqual(
            item.changed_fields['wholesale_price'],
            {'old': WHOLESALE, 'new': 320},
        )
        self.assertEqual(item.changed_fields['stock_qty'], {'old': None, 'new': 27})
        fulfillment = ProductFulfillment.objects.get(product=self.product)
        self.assertEqual(fulfillment.source, ProductFulfillment.SOURCE_MANUAL)
        self.assertEqual(fulfillment.external_id, '')
        self.assertIsNotNone(fulfillment.last_synced_at)

    def test_creates_min_qty_one_tier_and_keeps_wms_source(self):
        extra = Product.objects.create(
            title='Без опта',
            price=1000,
            seller_name=self.seller.name,
            seller_profile=self.seller,
            whatsapp_number='+77770001122',
            status='active',
            publish_to_sellers=True,
            city='Алматы',
            article='NO-TIER-1',
            slug='wh-upd-notier',
        )
        ProductFulfillment.objects.create(
            product=self.product,
            external_id='WMS-KEEP',
            source=ProductFulfillment.SOURCE_WMS,
        )
        payload = _xlsx_bytes([
            [extra.article, '', '', 410, ''],
            [self.product.article, '', '', '', 3],
        ])
        self._upload(payload)
        batch = CatalogImportBatch.objects.get(mode=CatalogImportBatch.MODE_DRY_RUN)
        self.client.post(self._apply_url(batch), {'confirm': '1'})
        self.assertTrue(
            ProductPriceTier.objects.filter(
                product=extra, min_qty=1, price=410, is_active=True
            ).exists()
        )
        fulfillment = ProductFulfillment.objects.get(product=self.product)
        self.assertEqual(fulfillment.source, ProductFulfillment.SOURCE_WMS)
        self.assertEqual(fulfillment.external_id, 'WMS-KEEP')
        self.assertIsNotNone(fulfillment.last_synced_at)

    def test_conflict_preview_cannot_apply_and_stale_blocks_all(self):
        payload = _xlsx_bytes([['MISSING', '', 1, 1, 1]])
        response = self._upload(payload)
        html = self.client.get(response.url).content.decode('utf-8')
        self.assertIn('Apply недоступна', html)
        self.assertNotIn('name="confirm"', html)
        batch = CatalogImportBatch.objects.get()
        blocked = self.client.post(self._apply_url(batch), {'confirm': '1'})
        self.assertEqual(blocked.status_code, 302)

        good = _xlsx_bytes([[self.product.article, '', 2200, '', '']])
        CatalogImportBatch.objects.all().delete()
        self._upload(good)
        batch = CatalogImportBatch.objects.get(mode=CatalogImportBatch.MODE_DRY_RUN)
        self.product.price = 2500
        self.product.save(update_fields=['price'])
        stale = self.client.post(self._apply_url(batch), {'confirm': '1'})
        self.assertEqual(stale.status_code, 302)
        self.product.refresh_from_db()
        self.assertEqual(self.product.price, 2500)
        self.assertFalse(
            CatalogImportBatch.objects.filter(mode=CatalogImportBatch.MODE_WRITE).exists()
        )

    def test_filename_is_not_used_as_path(self):
        payload = _xlsx_bytes([[self.product.article, '', 2200, '', '']])
        self._upload(payload, name=r'..\..\evil.xlsx')
        batch = CatalogImportBatch.objects.get()
        self.assertNotIn('..', batch.filename)
        self.assertEqual(batch.file_sha256, sha256_bytes(payload))
        self.assertEqual(len(batch.file_sha256), 64)
