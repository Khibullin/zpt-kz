from importlib import import_module

from django.apps import apps
from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook
from io import BytesIO

from catalog.models import Product, ProductPriceTier, SellerProfile, SellerWholesaleTerms
from catalog.wholesale import has_public_wholesale_offer, wholesale_products_qs
from marketing.models import (
    MarketingCampaign,
    MarketingCampaignMessage,
    MarketingCampaignRecipient,
    MarketingCampaignSendRun,
    MarketingWhatsAppTemplate,
)
from orders.models import Order, OrderItem


CLEANUP = import_module('catalog.migrations.0026_clean_ag_parts_wholesale_assortment')


def _make_seller(username, name, phone, **kwargs):
    user = User.objects.create_user(username=username, password='secret12345')
    defaults = {
        'user': user,
        'name': name,
        'phone': phone,
        'city': 'Алматы',
        'address': 'г. Алматы, ул. Тестовая, 1',
        'wholesale_enabled': True,
        'wholesale_min_order_qty': 10,
    }
    defaults.update(kwargs)
    return SellerProfile.objects.create(**defaults)


class AgPartsWholesaleCleanupTests(TestCase):
    def setUp(self):
        self.seller = _make_seller('ag-parts-owner', 'AG Parts', '77771360740')
        self.assertEqual(self.seller.slug, 'ag-parts')
        self.terms = SellerWholesaleTerms.objects.create(
            seller=self.seller,
            vat_mode=SellerWholesaleTerms.VAT_INCLUDED,
            prepayment_percent=100,
            pickup_city='Алматы',
            primary_carrier='DPD Kazakhstan',
        )
        self.whitelist_products = []
        for index, article in enumerate(sorted(CLEANUP.WHITELIST_ARTICLES)):
            product = self._product(
                title=f'Approved {article}',
                article=article,
                slug=f'ag-wl-{index}',
                price=2000 + index,
                stock_qty=11 + index,
                publish_to_sellers=True,
                publish_to_kaspi=(index == 0),
            )
            ProductPriceTier.objects.create(product=product, min_qty=1, price=400 + index)
            self.whitelist_products.append(product)

        self.air_filters = []
        for index, article in enumerate(sorted(CLEANUP.AIR_FILTER_ARTICLES)):
            product = self._product(
                title=f'Воздушный фильтр {article}',
                article=article,
                slug=f'ag-air-{index}',
                price=2310,
                publish_to_sellers=True,
            )
            ProductPriceTier.objects.create(product=product, min_qty=1, price=300)
            self.air_filters.append(product)

        self.brake = self._product(
            title='Тормозные колодки Haval',
            article='AG-BRAKE-1',
            slug='ag-brake-1',
            price=15000,
            publish_to_sellers=True,
        )
        ProductPriceTier.objects.create(product=self.brake, min_qty=1, price=9000)
        self.suspension = self._product(
            title='Сайлентблок подвески',
            article='AG-SUSP-1',
            slug='ag-susp-1',
            price=8900,
            publish_to_sellers=True,
        )
        ProductPriceTier.objects.create(product=self.suspension, min_qty=1, price=4000)

        self.canonical_spark = Product.objects.get(article='F4J163707010', seller_profile=self.seller)
        self.legacy_spark_a = self._product(
            title='Комплект свечей зажигания Iridium',
            article='F4J16-3707010',
            slug='ag-legacy-spark-a',
            price=9280,
            publish_to_sellers=False,
        )
        self.canonical_changan = Product.objects.get(article='D20T0120700', seller_profile=self.seller)
        self.legacy_spark_b = self._product(
            title='Комплект свечей зажигания Iridium 4 шт',
            article='D20T0120700,',
            slug='ag-legacy-spark-b',
            price=8649,
            publish_to_sellers=False,
        )
        self.suspect_a = self._product(
            title='Candidate A',
            article='CAND-100',
            slug='ag-cand-a',
            price=1111,
        )
        self.suspect_b = self._product(
            title='Candidate B',
            article='CAND-100,',
            slug='ag-cand-b',
            price=1112,
        )
        self.legacy_unbound = Product.objects.create(
            title='Legacy unbound cabin',
            price=1990,
            seller_name='AG Parts',
            seller_profile=None,
            whatsapp_number='+77771360740',
            status='active',
            publish_to_sellers=True,
            city='Алматы',
            article='LEGACY-UNBOUND-1',
            slug='ag-legacy-unbound',
        )
        ProductPriceTier.objects.create(product=self.legacy_unbound, min_qty=1, price=200)

        self.other = _make_seller('other-wh-owner', 'Other Shop', '77001112233')
        self.other_product = Product.objects.create(
            title='Чужой опт',
            price=5000,
            seller_name=self.other.name,
            seller_profile=self.other,
            whatsapp_number='+77001112233',
            status='active',
            publish_to_sellers=True,
            city='Алматы',
            article='OTHER-WH-1',
            slug='other-wh-1',
        )
        ProductPriceTier.objects.create(product=self.other_product, min_qty=1, price=2500)

        self.snapshot = self._snapshot_whitelist()
        self.terms_snapshot = self._terms_snapshot()
        self.template_snapshot = self._template_snapshot()

    def _product(self, **kwargs):
        defaults = {
            'seller_name': self.seller.name,
            'seller_profile': self.seller,
            'whatsapp_number': '+77771360740',
            'status': 'active',
            'city': 'Алматы',
            'publish_to_sellers': False,
            'publish_to_kaspi': False,
        }
        defaults.update(kwargs)
        return Product.objects.create(**defaults)

    def _snapshot_whitelist(self):
        data = {}
        for product in self.whitelist_products:
            product.refresh_from_db()
            tier = product.price_tiers.get(min_qty=1, is_active=True)
            data[product.article] = {
                'price': product.price,
                'stock_qty': product.stock_qty,
                'publish_to_sellers': product.publish_to_sellers,
                'publish_to_kaspi': product.publish_to_kaspi,
                'status': product.status,
                'seller_profile_id': product.seller_profile_id,
                'tier_price': tier.price,
                'tier_id': tier.pk,
            }
        return data

    def _terms_snapshot(self):
        self.terms.refresh_from_db()
        return {
            'vat_mode': self.terms.vat_mode,
            'prepayment_percent': self.terms.prepayment_percent,
            'pickup_city': self.terms.pickup_city,
            'primary_carrier': self.terms.primary_carrier,
        }

    def _template_snapshot(self):
        template = MarketingWhatsAppTemplate.objects.filter(
            meta_template_name='zpt_ag_parts_wholesale_v1',
            language_code='ru',
        ).first()
        if template is None:
            return None
        return {
            'meta_status': template.meta_status,
            'meta_template_id': template.meta_template_id,
            'body': template.body_text,
            'header': template.header_text,
        }

    def _run(self):
        CLEANUP.clean_ag_parts_wholesale_assortment(apps, None)
        CLEANUP.clean_ag_parts_wholesale_assortment(apps, None)

    def test_whitelist_stays_public_wholesale(self):
        self._run()
        wholesale = list(
            wholesale_products_qs(self.seller).order_by('article').values_list('article', flat=True)
        )
        self.assertEqual(len(wholesale), 26)
        self.assertEqual(set(wholesale), set(CLEANUP.WHITELIST_ARTICLES))
        for product in self.whitelist_products:
            product.refresh_from_db()
            snap = self.snapshot[product.article]
            self.assertEqual(product.price, snap['price'])
            self.assertEqual(product.stock_qty, snap['stock_qty'])
            self.assertTrue(product.publish_to_sellers)
            self.assertEqual(product.publish_to_kaspi, snap['publish_to_kaspi'])
            self.assertEqual(product.status, 'active')
            self.assertEqual(product.seller_profile_id, self.seller.pk)
            self.assertTrue(has_public_wholesale_offer(product, self.seller))
            tiers = list(product.price_tiers.filter(is_active=True))
            self.assertEqual(len(tiers), 1)
            self.assertEqual(tiers[0].min_qty, 1)
            self.assertEqual(tiers[0].price, snap['tier_price'])
            self.assertEqual(tiers[0].pk, snap['tier_id'])

    def test_air_filters_and_other_retail_lose_wholesale(self):
        self._run()
        for product in self.air_filters + [self.brake, self.suspension, self.legacy_unbound]:
            product.refresh_from_db()
            self.assertTrue(Product.objects.filter(pk=product.pk).exists())
            self.assertEqual(product.status, 'active')
            self.assertFalse(product.publish_to_sellers)
            self.assertFalse(product.price_tiers.exists())
            self.assertFalse(has_public_wholesale_offer(product, self.seller))

        self.brake.refresh_from_db()
        self.suspension.refresh_from_db()
        self.assertEqual(self.brake.price, 15000)
        self.assertEqual(self.suspension.price, 8900)

        html = self.client.get(
            reverse('catalog_list'),
            {'q': self.air_filters[0].article},
        ).content.decode('utf-8')
        self.assertIn(self.air_filters[0].article, html)
        self.assertNotIn('Купить оптом', html)
        self.assertNotIn('Опт от', html)

        other_html = self.client.get(
            reverse('catalog_list'),
            {'q': self.brake.article},
        ).content.decode('utf-8')
        self.assertIn(self.brake.title, other_html)
        self.assertNotIn('Купить оптом', other_html)

        self.other_product.refresh_from_db()
        self.assertTrue(self.other_product.publish_to_sellers)
        self.assertTrue(self.other_product.price_tiers.filter(is_active=True).exists())

    def test_known_spark_duplicates_are_removed(self):
        self._run()
        self.assertTrue(
            Product.objects.filter(
                pk=self.canonical_spark.pk,
                article='F4J163707010',
                status='active',
            ).exists()
        )
        self.assertTrue(
            Product.objects.filter(
                pk=self.canonical_changan.pk,
                article='D20T0120700',
                status='active',
            ).exists()
        )
        self.assertFalse(Product.objects.filter(pk=self.legacy_spark_a.pk).exists())
        self.assertFalse(Product.objects.filter(article='F4J16-3707010').exists())
        self.assertFalse(Product.objects.filter(pk=self.legacy_spark_b.pk).exists())
        self.assertFalse(Product.objects.filter(article='D20T0120700,').exists())

    def test_known_duplicate_with_order_is_hidden_not_deleted(self):
        order = Order.objects.create(
            customer_name='Buyer',
            customer_phone='77000000001',
            total_price=9280,
            delivery_method=Order.DELIVERY_PICKUP,
        )
        OrderItem.objects.create(
            order=order,
            product=self.legacy_spark_a,
            quantity=1,
            price_at_purchase=9280,
        )
        self._run()
        self.legacy_spark_a.refresh_from_db()
        self.assertEqual(self.legacy_spark_a.status, 'hidden')
        self.assertFalse(self.legacy_spark_a.publish_to_sellers)
        self.assertFalse(self.legacy_spark_a.price_tiers.exists())
        self.assertTrue(
            Product.objects.filter(pk=self.canonical_spark.pk, article='F4J163707010').exists()
        )

    def test_xlsx_contains_only_whitelist(self):
        self._run()
        response = self.client.get(
            reverse('public_seller_wholesale_price', kwargs={'slug': self.seller.slug})
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        articles = [
            row[0]
            for row in workbook['Прайс'].iter_rows(min_row=2, values_only=True)
            if row[0]
        ]
        self.assertEqual(len(articles), 26)
        self.assertEqual(set(articles), set(CLEANUP.WHITELIST_ARTICLES))

    def test_suspected_duplicates_are_reported_not_deleted(self):
        self._run()
        self.assertTrue(Product.objects.filter(pk=self.suspect_a.pk, status='active').exists())
        self.assertTrue(Product.objects.filter(pk=self.suspect_b.pk, status='active').exists())
        remaining = CLEANUP.ag_parts_products(Product, self.seller)
        suspected = CLEANUP.list_suspected_ag_parts_duplicates(remaining)
        articles = {item['article'] for item in suspected}
        self.assertIn('CAND-100', articles)
        self.assertIn('CAND-100,', articles)
        self.assertNotIn('F4J16-3707010', articles)
        self.assertNotIn('D20T0120700,', articles)

    def test_terms_marketing_and_min_order_untouched(self):
        self._run()
        self.seller.refresh_from_db()
        self.assertEqual(self.seller.wholesale_min_order_qty, 10)
        self.terms.refresh_from_db()
        self.assertEqual(self._terms_snapshot(), self.terms_snapshot)
        if self.template_snapshot is not None:
            template = MarketingWhatsAppTemplate.objects.get(
                meta_template_name='zpt_ag_parts_wholesale_v1',
                language_code='ru',
            )
            self.assertEqual(template.meta_status, self.template_snapshot['meta_status'])
            self.assertEqual(template.meta_template_id, self.template_snapshot['meta_template_id'])
            self.assertEqual(template.body_text, self.template_snapshot['body'])
            self.assertEqual(template.header_text, self.template_snapshot['header'])
        campaign = MarketingCampaign.objects.filter(
            name='AG Parts — запуск оптовой витрины — 08.2026',
        ).first()
        if campaign is not None:
            self.assertEqual(campaign.status, 'draft')
            self.assertEqual(
                MarketingCampaignRecipient.objects.filter(campaign=campaign).count(),
                0,
            )
            self.assertEqual(
                MarketingCampaignSendRun.objects.filter(campaign=campaign).count(),
                0,
            )
            self.assertEqual(
                MarketingCampaignMessage.objects.filter(send_run__campaign=campaign).count(),
                0,
            )

    def test_missing_seller_is_noop(self):
        self.seller.slug = 'ag-parts-other'
        self.seller.save(update_fields=['slug'])
        before = ProductPriceTier.objects.filter(product=self.brake).count()
        CLEANUP.clean_ag_parts_wholesale_assortment(apps, None)
        self.assertEqual(
            ProductPriceTier.objects.filter(product=self.brake).count(),
            before,
        )
        self.assertTrue(Product.objects.filter(pk=self.legacy_spark_a.pk).exists())
