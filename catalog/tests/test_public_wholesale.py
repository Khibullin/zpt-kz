from io import BytesIO

from django.contrib.auth.models import AnonymousUser, User
from django.db import connection
from django.test import RequestFactory, TestCase
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from catalog.applicability import extra_compatibility_text, public_card_fitment
from catalog.commercial import get_request_seller_profile, resolve_commercial_price
from catalog.models import Brand, Country, Product, ProductPriceTier, SellerProfile, SellerWholesaleTerms
from catalog.wholesale import (
    WHOLESALE_TYPE_CABIN,
    WHOLESALE_TYPE_OIL,
    WHOLESALE_TYPE_SPARK,
    public_wholesale_unit_price,
    safe_wholesale_filename_stem,
    wholesale_price_filename,
    wholesale_product_type,
)
from catalog.wholesale_export import XLSX_CONTENT_TYPE


RETAIL = 2500
WHOLESALE = 950


def _make_seller(username, name, phone, **kwargs):
    user = User.objects.create_user(username=username, password='secret12345')
    defaults = {
        'user': user,
        'name': name,
        'phone': phone,
        'city': 'Алматы',
        'address': 'г. Алматы, ул. Тестовая, 1',
        'wholesale_enabled': True,
        'wholesale_min_order_qty': 10,
    }
    defaults.update(kwargs)
    return SellerProfile.objects.create(**defaults)


class PublicWholesaleStorefrontTests(TestCase):
    def setUp(self):
        self.seller = _make_seller(
            'wholesale-owner',
            'AG Parts',
            '77771360740',
        )
        self.country = Country.objects.create(name='Китай WH')
        self.haval = Brand.objects.create(country=self.country, name='Haval')
        self.chery = Brand.objects.create(country=self.country, name='Chery')
        self.product = self._product(
            title='Салонный фильтр Haval Jolion',
            article='WH-CABIN-1',
            slug='wh-cabin-1',
        )
        self.product.selected_brands.add(self.haval)
        ProductPriceTier.objects.create(
            product=self.product,
            min_qty=1,
            price=WHOLESALE,
        )

    def _product(self, **kwargs):
        defaults = {
            'title': 'Оптовый товар',
            'price': RETAIL,
            'seller_name': self.seller.name,
            'seller_profile': self.seller,
            'whatsapp_number': '+77771360740',
            'status': 'active',
            'publish_to_sellers': True,
            'city': 'Алматы',
            'article': 'WH-1',
        }
        defaults.update(kwargs)
        return Product.objects.create(**defaults)

    def _url(self, seller=None):
        seller = seller or self.seller
        return reverse('public_seller_wholesale', kwargs={'slug': seller.slug})

    def _profile_url(self, seller=None):
        seller = seller or self.seller
        return reverse('public_seller_profile', kwargs={'slug': seller.slug})

    def _price_url(self, seller=None):
        seller = seller or self.seller
        return reverse('public_seller_wholesale_price', kwargs={'slug': seller.slug})

    def test_seller_wholesale_defaults(self):
        other = _make_seller(
            'plain-owner',
            'Plain Shop',
            '77770000099',
            wholesale_enabled=False,
        )
        self.assertFalse(other.wholesale_enabled)
        self.assertEqual(other.wholesale_min_order_qty, 10)

    def test_anonymous_sees_public_wholesale_price(self):
        response = self.client.get(self._url())
        self.assertEqual(response.status_code, 200)
        html = response.content.decode('utf-8')
        self.assertIn('Оптовая цена:', html)
        self.assertIn(str(WHOLESALE), html)
        self.assertIn('₸/шт', html)
        self.assertIn(self.product.title, html)
        self.assertIn(self.product.article, html)
        self.assertIn('Купить оптом', html)
        self.assertIn('product product-v2', html)
        self.assertIn('data-card-mode="wholesale"', html)
        self.assertIn('data-cart-mode="wholesale"', html)
        self.assertNotIn('wholesale-card', html)
        self.assertNotIn('wholesale-grid', html)
        self.assertNotIn('Есть оптовая цена', html)
        self.assertIn(
            'AG Parts — оптовые автозапчасти для китайских автомобилей',
            html,
        )
        self.assertIn(
            'Постоянные оптовые поставки для магазинов автозапчастей',
            html,
        )
        self.assertNotIn('автокомпонентов', html)
        self.assertNotIn('автокомпоненты', html.lower())
        self.assertIn(
            'Минимальный оптовый заказ — 10 единиц в ассортименте',
            html,
        )
        lowered = html.lower()
        for banned in ('скидка', 'распродажа', 'акция'):
            self.assertNotIn(banned, lowered)

    def test_anonymous_retail_catalog_price_unchanged(self):
        factory = RequestFactory()
        request = factory.get('/market/')
        request.user = AnonymousUser()
        quote = resolve_commercial_price(
            self.product,
            10,
            seller_profile=get_request_seller_profile(request),
        )
        self.assertEqual(quote.unit_price, RETAIL)
        self.assertEqual(public_wholesale_unit_price(self.product), WHOLESALE)

    def test_publish_to_sellers_false_excluded(self):
        hidden_flag = self._product(
            title='Не для продавцов фильтр',
            article='WH-NOPUB',
            slug='wh-nopub',
            publish_to_sellers=False,
        )
        ProductPriceTier.objects.create(product=hidden_flag, min_qty=1, price=700)
        html = self.client.get(self._url()).content.decode('utf-8')
        self.assertNotIn(hidden_flag.title, html)
        self.assertNotIn(hidden_flag.article, html)

    def test_hidden_product_excluded(self):
        hidden = self._product(
            title='Скрытый салонный фильтр',
            article='WH-HIDDEN',
            slug='wh-hidden',
            status='hidden',
        )
        ProductPriceTier.objects.create(product=hidden, min_qty=1, price=700)
        html = self.client.get(self._url()).content.decode('utf-8')
        self.assertNotIn(hidden.title, html)

    def test_product_without_active_tier_excluded(self):
        no_tier = self._product(
            title='Без оптовой цены фильтр',
            article='WH-NOTIER',
            slug='wh-notier',
        )
        inactive = self._product(
            title='Неактивный тариф фильтр',
            article='WH-INACTIVE',
            slug='wh-inactive',
        )
        ProductPriceTier.objects.create(
            product=inactive,
            min_qty=1,
            price=700,
            is_active=False,
        )
        html = self.client.get(self._url()).content.decode('utf-8')
        self.assertNotIn(no_tier.title, html)
        self.assertNotIn(inactive.title, html)

    def test_wholesale_disabled_storefront_is_404(self):
        self.seller.wholesale_enabled = False
        self.seller.save(update_fields=['wholesale_enabled'])
        response = self.client.get(self._url())
        self.assertEqual(response.status_code, 404)

    def test_public_profile_shows_wholesale_cta(self):
        url = reverse('public_seller_profile', kwargs={'slug': self.seller.slug})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        html = response.content.decode('utf-8')
        self.assertIn('Оптовые цены', html)
        self.assertIn(self._url(), html)

    def test_public_profile_hides_cta_when_disabled(self):
        self.seller.wholesale_enabled = False
        self.seller.save(update_fields=['wholesale_enabled'])
        url = reverse('public_seller_profile', kwargs={'slug': self.seller.slug})
        html = self.client.get(url).content.decode('utf-8')
        self.assertNotIn('Оптовые цены', html)

    def test_type_and_brand_filters(self):
        oil = self._product(
            title='Масляный фильтр Chery Tiggo',
            article='WH-OIL-1',
            slug='wh-oil-1',
        )
        oil.selected_brands.add(self.chery)
        ProductPriceTier.objects.create(product=oil, min_qty=1, price=880)
        spark = self._product(
            title='Свеча зажигания Haval',
            article='WH-SPARK-1',
            slug='wh-spark-1',
        )
        spark.selected_brands.add(self.haval)
        ProductPriceTier.objects.create(product=spark, min_qty=1, price=420)

        self.assertEqual(wholesale_product_type(self.product), WHOLESALE_TYPE_CABIN)
        self.assertEqual(wholesale_product_type(oil), WHOLESALE_TYPE_OIL)
        self.assertEqual(wholesale_product_type(spark), WHOLESALE_TYPE_SPARK)

        cabin_html = self.client.get(
            self._url(),
            {'type': WHOLESALE_TYPE_CABIN},
        ).content.decode('utf-8')
        self.assertIn(self.product.title, cabin_html)
        self.assertNotIn(oil.title, cabin_html)

        brand_html = self.client.get(
            self._url(),
            {'brand': self.chery.id},
        ).content.decode('utf-8')
        self.assertIn(oil.title, brand_html)
        self.assertNotIn(self.product.title, brand_html)

        search_html = self.client.get(
            self._url(),
            {'q': 'Jolion'},
        ).content.decode('utf-8')
        self.assertIn(self.product.title, search_html)
        self.assertNotIn(oil.title, search_html)

    def _detail_url(self, product=None):
        product = product or self.product
        return reverse('product_detail', kwargs={'slug': product.slug})

    def test_anonymous_product_detail_shows_wholesale_block(self):
        response = self.client.get(self._detail_url())
        self.assertEqual(response.status_code, 200)
        html = response.content.decode('utf-8').replace('\xa0', ' ')
        self.assertIn('Розничная цена:', html)
        self.assertIn('2 500', html)
        self.assertIn('Оптовая цена:', html)
        self.assertIn(str(WHOLESALE), html)
        self.assertIn('Минимальный оптовый заказ — 10 единиц', html)
        self.assertIn('Купить оптом', html)
        self.assertIn('Купить в розницу', html)
        self.assertIn('Смотреть весь оптовый ассортимент продавца', html)
        self.assertIn(self._url(), html)
        self.assertIn('Наличие уточняется', html)
        self.assertNotIn('с НДС', html)
        self.assertNotIn('без НДС', html)
        self.assertNotIn('100% предоплата', html)
        self.assertIn('Условия оптовой покупки', html)
        self.assertIn('#wholesale-terms', html)
        lowered = html.lower()
        for banned in ('скидка', 'распродажа', 'акция'):
            self.assertNotIn(banned, lowered)

    def test_ineligible_product_detail_has_no_wholesale_block(self):
        ineligible = self._product(
            title='Обычный розничный товар',
            article='WH-RETAIL-ONLY',
            slug='wh-retail-only',
            publish_to_sellers=False,
        )
        response = self.client.get(self._detail_url(ineligible))
        html = response.content.decode('utf-8')
        self.assertNotIn('Оптовая цена:', html)
        self.assertNotIn('Купить оптом', html)
        self.assertNotIn('Есть оптовая цена', html)
        self.assertNotIn('с НДС', html)
        self.assertNotIn('Условия оптовой покупки', html)

    def test_public_stock_null_unknown_and_orderable(self):
        self.assertIsNone(self.product.stock_qty)
        html = self.client.get(self._url()).content.decode('utf-8')
        self.assertIn('Наличие уточняется', html)
        self.assertIn('Купить оптом', html)
        self.assertNotIn('Нет в наличии', html)
        self.assertNotIn('Осталось всего', html)
        detail = self.client.get(self._detail_url()).content.decode('utf-8')
        self.assertIn('Наличие уточняется', detail)
        self.assertIn('Купить оптом', detail)

    def test_public_stock_zero_visible_but_not_buyable(self):
        self.product.stock_qty = 0
        self.product.save(update_fields=['stock_qty'])
        html = self.client.get(self._url()).content.decode('utf-8')
        self.assertIn('Нет в наличии', html)
        self.assertNotIn('data-wholesale-add', html)
        self.assertNotIn('data-cart-add', html)
        detail = self.client.get(self._detail_url()).content.decode('utf-8')
        self.assertIn('Нет в наличии', detail)
        self.assertIn('Купить оптом недоступно', detail)

    def test_public_stock_positive_shows_count(self):
        self.product.stock_qty = 27
        self.product.save(update_fields=['stock_qty'])
        html = self.client.get(self._url()).content.decode('utf-8')
        self.assertIn('В наличии: 27 шт.', html)
        self.assertIn('Купить оптом', html)
        self.assertIn('data-max-qty="27"', html)

    def test_price_xlsx_stock_values(self):
        self.product.stock_qty = 0
        self.product.save(update_fields=['stock_qty'])
        extra = self._product(title='In stock', article='WH-STK-1', slug='wh-stk-1')
        extra.stock_qty = 4
        extra.save(update_fields=['stock_qty'])
        ProductPriceTier.objects.create(product=extra, min_qty=1, price=400)
        workbook = load_workbook(BytesIO(self.client.get(self._price_url()).content))
        by_article = {
            row[0]: row
            for row in workbook['Прайс'].iter_rows(min_row=2, values_only=True)
            if row[0]
        }
        self.assertEqual(by_article[self.product.article][7], 0)
        self.assertEqual(by_article[extra.article][7], 4)

    def test_catalog_shows_wholesale_price_not_badge(self):
        catalog = self.client.get(reverse('catalog_list'), {'q': self.product.article})
        html = catalog.content.decode('utf-8').replace('\xa0', ' ')
        self.assertIn('product product-v2', html)
        self.assertNotIn('Есть оптовая цена', html)
        self.assertIn('Оптовая цена:', html)
        self.assertIn(f'{WHOLESALE} ₸/шт', html)
        self.assertIn('2 500', html)
        self.assertIn('data-card-mode="retail"', html)
        self.assertIn('data-cart-mode="retail"', html)

    def test_catalog_hides_badge_when_not_eligible(self):
        no_publish = self._product(
            title='Без публикации продавцам',
            article='WH-NOBADGE-1',
            slug='wh-nobadge-1',
            publish_to_sellers=False,
        )
        ProductPriceTier.objects.create(product=no_publish, min_qty=1, price=400)
        no_tier = self._product(
            title='Без активного тарифа',
            article='WH-NOBADGE-2',
            slug='wh-nobadge-2',
        )
        disabled_seller = _make_seller(
            'disabled-wh',
            'No Wholesale Shop',
            '77770000088',
            wholesale_enabled=False,
        )
        disabled_product = Product.objects.create(
            title='Товар выключенного опта',
            price=RETAIL,
            seller_name=disabled_seller.name,
            seller_profile=disabled_seller,
            whatsapp_number='+77770000088',
            status='active',
            publish_to_sellers=True,
            city='Алматы',
            article='WH-NOBADGE-3',
            slug='wh-nobadge-3',
        )
        ProductPriceTier.objects.create(product=disabled_product, min_qty=1, price=400)

        for article, title in (
            (no_publish.article, no_publish.title),
            (no_tier.article, no_tier.title),
            (disabled_product.article, disabled_product.title),
        ):
            html = self.client.get(
                reverse('catalog_list'),
                {'q': article},
            ).content.decode('utf-8')
            self.assertIn(title, html)
            self.assertNotIn('Есть оптовая цена', html)
            self.assertNotIn('Оптовая цена:', html)

    def test_catalog_wholesale_badge_query_count_stable(self):
        for index in range(4):
            extra = self._product(
                title=f'Салонный фильтр extra {index}',
                article=f'WH-N1-{index}',
                slug=f'wh-n1-{index}',
            )
            ProductPriceTier.objects.create(product=extra, min_qty=1, price=800)
        url = reverse('catalog_list')
        params = {'q': 'Салонный фильтр'}
        with CaptureQueriesContext(connection) as first:
            self.client.get(url, params)
        baseline = len(first.captured_queries)
        for index in range(4, 8):
            extra = self._product(
                title=f'Салонный фильтр extra {index}',
                article=f'WH-N1-{index}',
                slug=f'wh-n1-{index}',
            )
            ProductPriceTier.objects.create(product=extra, min_qty=1, price=800)
        with CaptureQueriesContext(connection) as second:
            response = self.client.get(url, params)
        self.assertLessEqual(len(second.captured_queries), baseline + 1)
        self.assertContains(response, 'Оптовая цена:')
        self.assertNotContains(response, 'Есть оптовая цена')

    def test_product_detail_captures_utm_without_empty_overwrite(self):
        from orders.constants import SESSION_UTM_KEY

        self.client.get(
            self._detail_url(),
            {
                'utm_source': 'whatsapp',
                'utm_medium': 'marketing',
                'utm_campaign': 'launch',
            },
        )
        self.client.get(self._detail_url())
        stored = self.client.session[SESSION_UTM_KEY]
        self.assertEqual(stored['utm_source'], 'whatsapp')
        self.assertEqual(stored['utm_campaign'], 'launch')

    def test_catalog_retail_only_has_no_wholesale_price_or_unknown_stock(self):
        retail_only = self._product(
            title='Обычный розничный фильтр',
            article='WH-RETAIL-CARD',
            slug='wh-retail-card',
            publish_to_sellers=False,
        )
        html = self.client.get(
            reverse('catalog_list'),
            {'q': retail_only.article},
        ).content.decode('utf-8').replace('\xa0', ' ')
        self.assertIn(retail_only.title, html)
        self.assertIn('2 500', html)
        self.assertNotIn('Оптовая цена:', html)
        self.assertNotIn('Есть оптовая цена', html)
        self.assertNotIn('Наличие уточняется', html)

    def test_wholesale_storefront_shows_retail_and_tier_price(self):
        html = self.client.get(self._url()).content.decode('utf-8').replace('\xa0', ' ')
        self.assertIn('2 500', html)
        self.assertIn('Оптовая цена:', html)
        self.assertIn(f'{WHOLESALE} ₸/шт', html)
        self.assertIn('qty-input', html)
        self.assertIn('data-qty-minus', html)
        self.assertIn('data-qty-plus', html)
        self.assertIn('WhatsApp продавцу', html)
        self.assertIn('Подробнее', html)

    def test_public_card_hides_research_notes(self):
        note = 'у поставщиков, в справочнике ZPT нет. Списки отвергнуты'
        self.product.compatibility = note
        self.product.save(update_fields=['compatibility'])
        self.assertIn('у поставщиков', extra_compatibility_text(self.product))
        self.assertNotIn('у поставщиков', public_card_fitment(self.product))
        catalog = self.client.get(
            reverse('catalog_list'),
            {'q': self.product.article},
        ).content.decode('utf-8')
        wholesale = self.client.get(self._url()).content.decode('utf-8')
        profile = self.client.get(self._profile_url()).content.decode('utf-8')
        for html in (catalog, wholesale, profile):
            self.assertIn(self.product.title, html)
            self.assertNotIn('у поставщиков, в справочнике ZPT нет', html)
            self.assertNotIn('Списки отвергнуты', html)
            self.assertNotIn('не подтверждено', html)

    def test_catalog_unknown_stock_hidden_unless_wholesale(self):
        self.assertIsNone(self.product.stock_qty)
        html = self.client.get(
            reverse('catalog_list'),
            {'q': self.product.article},
        ).content.decode('utf-8')
        self.assertIn('Наличие уточняется', html)

    def test_shared_product_card_on_catalog_profile_and_wholesale(self):
        catalog = self.client.get(
            reverse('catalog_list'),
            {'q': self.product.article},
        ).content.decode('utf-8')
        profile = self.client.get(self._profile_url()).content.decode('utf-8')
        wholesale = self.client.get(self._url()).content.decode('utf-8')

        for html in (catalog, profile, wholesale):
            self.assertIn('product product-v2', html)
            self.assertNotIn('seller-product-card', html)
            self.assertNotIn('wholesale-card', html)
            self.assertNotIn('Есть оптовая цена', html)

        self.assertIn('data-card-mode="retail"', catalog)
        self.assertIn('data-cart-mode="retail"', catalog)
        self.assertIn('data-card-mode="retail"', profile)
        self.assertIn('data-cart-mode="retail"', profile)
        self.assertIn('data-card-mode="wholesale"', wholesale)
        self.assertIn('data-cart-mode="wholesale"', wholesale)
        self.assertIn('WhatsApp продавцу', profile)
        self.assertIn('Подробнее', profile)
        self.assertIn('qty-input', profile)

    def test_seller_profile_shows_exact_wholesale_price(self):
        html = self.client.get(
            self._profile_url(),
            {'q_seller': self.product.article},
        ).content.decode('utf-8').replace('\xa0', ' ')
        self.assertIn(self.product.title, html)
        self.assertIn('2 500', html)
        self.assertIn('Оптовая цена:', html)
        self.assertIn(f'{WHOLESALE} ₸/шт', html)
        self.assertNotIn('seller-product-card', html)

    def test_seller_profile_retail_only_hides_wholesale_price(self):
        retail_only = self._product(
            title='Розничный товар профиля',
            article='WH-PROFILE-RETAIL',
            slug='wh-profile-retail',
            publish_to_sellers=False,
        )
        html = self.client.get(
            self._profile_url(),
            {'q_seller': retail_only.article},
        ).content.decode('utf-8').replace('\xa0', ' ')
        self.assertIn(retail_only.title, html)
        self.assertIn('2 500', html)
        self.assertNotIn('Оптовая цена:', html)

    def test_seller_profile_wholesale_flags_query_count_stable(self):
        for index in range(4):
            extra = self._product(
                title=f'Профиль фильтр extra {index}',
                article=f'WH-PRF-{index}',
                slug=f'wh-prf-{index}',
            )
            ProductPriceTier.objects.create(product=extra, min_qty=1, price=800)
        url = self._profile_url()
        with CaptureQueriesContext(connection) as first:
            self.client.get(url)
        baseline = len(first.captured_queries)
        for index in range(4, 8):
            extra = self._product(
                title=f'Профиль фильтр extra {index}',
                article=f'WH-PRF-{index}',
                slug=f'wh-prf-{index}',
            )
            ProductPriceTier.objects.create(product=extra, min_qty=1, price=800)
        with CaptureQueriesContext(connection) as second:
            response = self.client.get(url)
        self.assertLessEqual(len(second.captured_queries), baseline + 1)
        self.assertContains(response, 'Оптовая цена:')


def _configured_terms(seller, **kwargs):
    defaults = {
        'vat_mode': SellerWholesaleTerms.VAT_INCLUDED,
        'prepayment_percent': 100,
        'confirm_stock_before_payment': True,
        'provides_invoice': True,
        'provides_waybill': True,
        'provides_esf': True,
        'pickup_enabled': True,
        'pickup_city': 'Алматы',
        'delivery_kz_enabled': True,
        'delivery_payer': SellerWholesaleTerms.DELIVERY_PAYER_BUYER,
        'primary_carrier': 'DPD Kazakhstan',
        'primary_carrier_service': 'DPD OPTIMUM',
        'primary_carrier_url': 'https://dpd.kz/',
        'other_carrier_allowed': True,
        'stock_note': 'Наличие подтверждается перед оплатой.',
    }
    defaults.update(kwargs)
    return SellerWholesaleTerms.objects.create(seller=seller, **defaults)


class SellerWholesaleTermsPublicTests(PublicWholesaleStorefrontTests):
    def _price_url(self, seller=None):
        seller = seller or self.seller
        return reverse('public_seller_wholesale_price', kwargs={'slug': seller.slug})

    def test_neutral_seller_without_terms_has_no_invented_rules(self):
        html = self.client.get(self._url()).content.decode('utf-8')
        self.assertIn('Условия оптовой покупки', html)
        self.assertIn('Минимальный заказ — 10 единиц в ассортименте', html)
        lowered = html.lower()
        self.assertNotIn('с ндс', lowered)
        self.assertNotIn('без ндс', lowered)
        self.assertNotIn('100% предоплата', html)
        self.assertNotIn('эсф', lowered)
        self.assertNotIn('dpd', lowered)
        self.assertNotIn('накладная', lowered)

    def test_vat_included_shows_on_storefront_and_product(self):
        _configured_terms(self.seller)
        storefront = self.client.get(self._url()).content.decode('utf-8')
        self.assertIn('Все оптовые цены указаны с НДС', storefront)
        self.assertIn('100% предоплата после подтверждения наличия', storefront)
        self.assertIn('Счет, накладная и ЭСФ', storefront)
        self.assertIn('Самовывоз — Алматы', storefront)
        self.assertIn('Доставка по Казахстану — DPD Kazakhstan', storefront)
        self.assertIn('Стоимость доставки оплачивает покупатель', storefront)
        self.assertIn('Другая транспортная компания — по согласованию', storefront)
        self.assertIn('Скачать оптовый прайс', storefront)
        self.assertIn(self._price_url(), storefront)
        detail = self.client.get(self._detail_url()).content.decode('utf-8').replace('\xa0', ' ')
        self.assertIn('с НДС', detail)
        self.assertIn('100% предоплата после подтверждения наличия.', detail)
        self.assertIn('#wholesale-terms', detail)
        self.assertNotIn('Счет, накладная и ЭСФ', detail)

    def test_vat_excluded_shows_correctly(self):
        _configured_terms(
            self.seller,
            vat_mode=SellerWholesaleTerms.VAT_EXCLUDED,
            prepayment_percent=None,
            confirm_stock_before_payment=False,
            provides_invoice=False,
            provides_waybill=False,
            provides_esf=False,
            pickup_enabled=False,
            pickup_city='',
            delivery_kz_enabled=False,
            other_carrier_allowed=False,
            stock_note='',
        )
        storefront = self.client.get(self._url()).content.decode('utf-8')
        self.assertIn('Оптовые цены указаны без НДС', storefront)
        self.assertNotIn('с НДС', storefront)
        self.assertNotIn('100% предоплата', storefront)
        self.assertNotIn('DPD', storefront)
        detail = self.client.get(self._detail_url()).content.decode('utf-8')
        self.assertIn('без НДС', detail)
        self.assertNotIn('· с НДС', detail)

    def test_price_xlsx_anonymous_enabled_storefront(self):
        _configured_terms(self.seller)
        self.product.cost_price = 87654321
        self.product.save(update_fields=['cost_price'])
        response = self.client.get(self._price_url())
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], XLSX_CONTENT_TYPE)
        filename = wholesale_price_filename(self.seller, day=timezone.localdate())
        self.assertEqual(filename, f'AG_Parts_wholesale_price_{timezone.localdate():%Y-%m-%d}.xlsx')
        self.assertIn(filename, response['Content-Disposition'])
        self.assertIn('attachment', response['Content-Disposition'])
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames[0], 'Прайс')
        self.assertIn('Условия', workbook.sheetnames)
        prices = workbook['Прайс']
        headers = [cell.value for cell in next(prices.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(headers[0], 'Артикул')
        self.assertEqual(headers[4], 'Розничная цена, ₸')
        self.assertEqual(headers[5], 'Оптовая цена, ₸ с НДС')
        self.assertNotIn('cost_price', headers)
        values = []
        for row in prices.iter_rows(min_row=2, values_only=True):
            values.append(row)
        self.assertEqual(len(values), 1)
        self.assertEqual(values[0][0], self.product.article)
        self.assertEqual(values[0][4], RETAIL)
        self.assertEqual(values[0][5], WHOLESALE)
        self.assertEqual(values[0][7], 'Уточняется при заказе')
        blob = ' '.join(str(item) for item in headers + list(values[0]))
        self.assertNotIn('87654321', blob)
        self.assertNotIn('cost_price', blob.lower())
        terms_text = ' '.join(
            str(value)
            for row in workbook['Условия'].iter_rows(values_only=True)
            for value in row
            if value
        )
        self.assertIn('с НДС', terms_text)
        self.assertIn('100% предоплата', terms_text)

    def test_price_xlsx_excludes_ineligible_products(self):
        other = _make_seller('xlsx-other', 'Other XLSX', '77770000011')
        other_product = Product.objects.create(
            title='Чужой опт',
            price=111,
            seller_name=other.name,
            seller_profile=other,
            whatsapp_number='+77770000011',
            status='active',
            publish_to_sellers=True,
            city='Алматы',
            article='WH-OTHER-X',
            slug='wh-other-x',
        )
        ProductPriceTier.objects.create(product=other_product, min_qty=1, price=50)
        hidden = self._product(
            title='Скрытый',
            article='WH-HID-X',
            slug='wh-hid-x',
            status='hidden',
        )
        ProductPriceTier.objects.create(product=hidden, min_qty=1, price=40)
        unpublished = self._product(
            title='Не для продавцов',
            article='WH-NP-X',
            slug='wh-np-x',
            publish_to_sellers=False,
        )
        ProductPriceTier.objects.create(product=unpublished, min_qty=1, price=40)
        no_tier = self._product(title='Без тарифа', article='WH-NT-X', slug='wh-nt-x')
        response = self.client.get(self._price_url())
        workbook = load_workbook(BytesIO(response.content))
        articles = [
            row[0]
            for row in workbook['Прайс'].iter_rows(min_row=2, values_only=True)
            if row[0]
        ]
        self.assertEqual(articles, [self.product.article])
        self.assertNotIn(other_product.article, articles)
        self.assertNotIn(hidden.article, articles)
        self.assertNotIn(unpublished.article, articles)
        self.assertNotIn(no_tier.article, articles)

    def test_price_xlsx_disabled_seller_is_404(self):
        self.seller.wholesale_enabled = False
        self.seller.save(update_fields=['wholesale_enabled'])
        self.assertEqual(self.client.get(self._price_url()).status_code, 404)

    def test_price_xlsx_enabled_without_products_is_200(self):
        empty = _make_seller('xlsx-empty', 'Empty Opt', '77770000022')
        response = self.client.get(
            reverse('public_seller_wholesale_price', kwargs={'slug': empty.slug})
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook['Прайс'].iter_rows(min_row=2, values_only=True))
        self.assertEqual(rows, [])

    def test_filename_is_safe(self):
        self.assertEqual(safe_wholesale_filename_stem('AG Parts'), 'AG_Parts')
        self.assertEqual(safe_wholesale_filename_stem('Foo/Bar*.xlsx'), 'Foo_Bar_xlsx')
        self.assertEqual(safe_wholesale_filename_stem('Опт Детали'), 'wholesale')
        ugly = _make_seller('xlsx-ugly', 'Foo/Bar*.xlsx', '77770000033')
        name = wholesale_price_filename(ugly, day=timezone.localdate())
        self.assertTrue(name.startswith('Foo_Bar_xlsx_wholesale_price_'))
        self.assertTrue(name.endswith('.xlsx'))
        self.assertNotIn('/', name)
        self.assertNotIn('*', name)

    def test_price_xlsx_query_count_not_n_plus_one(self):
        for index in range(4):
            extra = self._product(
                title=f'Салонный фильтр xlsx {index}',
                article=f'WH-XLSX-{index}',
                slug=f'wh-xlsx-{index}',
            )
            ProductPriceTier.objects.create(product=extra, min_qty=1, price=800)
        url = self._price_url()
        with CaptureQueriesContext(connection) as first:
            first_response = self.client.get(url)
        self.assertEqual(first_response.status_code, 200)
        baseline = len(first.captured_queries)
        for index in range(4, 8):
            extra = self._product(
                title=f'Салонный фильтр xlsx {index}',
                article=f'WH-XLSX-{index}',
                slug=f'wh-xlsx-{index}',
            )
            ProductPriceTier.objects.create(product=extra, min_qty=1, price=800)
        with CaptureQueriesContext(connection) as second:
            second_response = self.client.get(url)
        self.assertEqual(second_response.status_code, 200)
        self.assertLessEqual(len(second.captured_queries), baseline + 1)
        workbook = load_workbook(BytesIO(second_response.content))
        articles = [
            row[0]
            for row in workbook['Прайс'].iter_rows(min_row=2, values_only=True)
            if row[0]
        ]
        self.assertGreaterEqual(len(articles), 9)


class AgPartsDescriptionMigrationTests(TestCase):
    def test_ag_parts_description_updated_and_missing_seller_is_noop(self):
        from importlib import import_module

        from django.apps import apps as django_apps

        migration = import_module(
            'catalog.migrations.0025_ag_parts_description_avtozapchasti'
        )

        migration.update_ag_parts_description(django_apps, None)

        other = _make_seller(
            'other-desc-owner',
            'Other Shop',
            '77001119988',
            slug='other-shop-desc',
            description='Keep me',
        )
        seller = _make_seller(
            'ag-parts-desc-owner',
            'AG Parts',
            '77771360999',
            slug='ag-parts',
            description=(
                'AG Parts — представитель группы китайских производителей '
                'автокомпонентов.'
            ),
        )
        migration.update_ag_parts_description(django_apps, None)
        seller.refresh_from_db()
        other.refresh_from_db()
        self.assertEqual(seller.description, migration.NEW_DESCRIPTION)
        self.assertEqual(other.description, 'Keep me')

        migration.update_ag_parts_description(django_apps, None)
        seller.refresh_from_db()
        self.assertEqual(seller.description, migration.NEW_DESCRIPTION)


