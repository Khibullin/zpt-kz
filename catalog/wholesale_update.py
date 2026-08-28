"""Seller-scoped wholesale price/stock XLSX update (admin-only).

Does not replace import_ag_parts. Preview and apply persist CatalogImportBatch
rows with source=wholesale_update.
"""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import PurePosixPath

from django import forms
from django.db import transaction
from django.db.models import Prefetch
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from catalog.models import (
    CatalogImportBatch,
    CatalogImportItem,
    Product,
    ProductFulfillment,
    ProductPriceTier,
)
from catalog.wholesale_export import XLSX_CONTENT_TYPE

UNCHANGED = object()
MAX_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_DATA_ROWS = 10_000
WHOLESALE_GTE_RETAIL_WARNING = (
    'Оптовая цена не ниже розничной — проверьте значение.'
)
STALE_APPLY_MESSAGE = (
    'Данные изменились после предварительной проверки. '
    'Создайте новый preview.'
)

HEADER_ARTICLE = 'артикул'
HEADER_TITLE = 'наименование'
HEADER_RETAIL = 'розничная цена'
HEADER_WHOLESALE = 'оптовая цена'
HEADER_STOCK = 'остаток'

_HEADER_ALIASES = {
    HEADER_ARTICLE: {'артикул', 'article', 'sku'},
    HEADER_TITLE: {'наименование', 'название', 'title'},
    HEADER_RETAIL: {'розничная цена', 'розница', 'retail'},
    HEADER_WHOLESALE: {'оптовая цена', 'опт', 'wholesale'},
    HEADER_STOCK: {'остаток', 'наличие', 'stock'},
}
_HEADER_NOISE = re.compile(r'[₸.,/]+')

DOWNLOAD_HEADERS = [
    'Артикул',
    'Наименование',
    'Розничная цена',
    'Оптовая цена',
    'Остаток',
]


class WholesaleUpdateUploadForm(forms.Form):
    file = forms.FileField(label='Файл XLSX')

    def clean_file(self):
        uploaded = self.cleaned_data['file']
        name = str(getattr(uploaded, 'name', '') or '')
        if not name.lower().endswith('.xlsx'):
            raise forms.ValidationError('Нужен файл в формате .xlsx.')
        size = getattr(uploaded, 'size', None)
        if size is not None and size > MAX_UPLOAD_BYTES:
            raise forms.ValidationError('Файл больше 5 МБ.')
        header = uploaded.read(4)
        uploaded.seek(0)
        if header != b'PK\x03\x04' and header[:2] != b'PK':
            raise forms.ValidationError('Файл не похож на XLSX.')
        if size is None:
            payload = uploaded.read()
            uploaded.seek(0)
            if len(payload) > MAX_UPLOAD_BYTES:
                raise forms.ValidationError('Файл больше 5 МБ.')
        return uploaded


class WholesaleUpdateError(Exception):
    pass


class WholesaleUpdateStaleError(WholesaleUpdateError):
    pass


class WholesaleUpdateBlockedError(WholesaleUpdateError):
    pass


@dataclass
class WholesaleUpdateRow:
    row_number: int
    article: str
    action: str
    product: Product | None = None
    errors: list = field(default_factory=list)
    warnings: list = field(default_factory=list)
    changed_fields: dict = field(default_factory=dict)
    baseline: dict = field(default_factory=dict)
    retail_new: object = UNCHANGED
    wholesale_new: object = UNCHANGED
    stock_new: object = UNCHANGED

    @property
    def will_change(self):
        return self.action == CatalogImportItem.ACTION_UPDATED


def _normalize_header(value):
    text = ' '.join(str(value or '').strip().casefold().split())
    text = text.replace('₸', ' ')
    text = _HEADER_NOISE.sub(' ', text)
    return ' '.join(text.split())


def _header_key(value):
    normalized = _normalize_header(value)
    for key, aliases in _HEADER_ALIASES.items():
        if normalized == key or normalized in aliases:
            return key
        if normalized.startswith(key):
            return key
    return ''


def min_qty_one_wholesale_price(product):
    cached = getattr(product, '_prefetched_objects_cache', {})
    if 'price_tiers' in cached:
        for tier in product.price_tiers.all():
            if getattr(tier, 'is_active', False) and int(tier.min_qty) == 1:
                return int(tier.price)
        return None
    tier = product.price_tiers.filter(is_active=True, min_qty=1).first()
    return int(tier.price) if tier is not None else None


def seller_update_products_qs(seller):
    return (
        Product.objects.owned_by_seller(seller)
        .select_related('seller_profile', 'brand', 'category')
        .prefetch_related(
            Prefetch(
                'price_tiers',
                queryset=ProductPriceTier.objects.filter(is_active=True).order_by(
                    'min_qty', 'id'
                ),
            ),
        )
        .order_by('article', 'id')
    )


def wholesale_update_filename(seller, day=None):
    day = day or timezone.localdate()
    slug = re.sub(r'[^A-Za-z0-9_-]+', '_', (getattr(seller, 'slug', '') or 'seller'))
    slug = slug.strip('_') or 'seller'
    return f'{slug}_prices_stock_{day:%Y-%m-%d}.xlsx'


def build_wholesale_update_workbook(seller):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Цены и остатки'
    sheet.append(DOWNLOAD_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for product in seller_update_products_qs(seller):
        stock = product.stock_qty
        sheet.append([
            product.article or '',
            product.title or '',
            int(product.price) if product.price is not None else '',
            min_qty_one_wholesale_price(product) or '',
            int(stock) if stock is not None else '',
        ])
    for index, width in enumerate((18, 42, 18, 18, 14), start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    return workbook


def wholesale_update_xlsx_bytes(seller):
    buffer = BytesIO()
    build_wholesale_update_workbook(seller).save(buffer)
    return buffer.getvalue()


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def safe_upload_filename(name: str) -> str:
    base = PurePosixPath(str(name or '').replace('\\', '/')).name
    base = re.sub(r'[^\w.\-]+', '_', base, flags=re.UNICODE).strip('._') or 'upload.xlsx'
    if not base.lower().endswith('.xlsx'):
        base = f'{base}.xlsx'
    return base[:255]


def _cell_is_formula(cell):
    if cell is None:
        return False
    if getattr(cell, 'data_type', '') == 'f':
        return True
    value = getattr(cell, 'value', None)
    return isinstance(value, str) and value.startswith('=')


def _is_empty_cell(value):
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _parse_positive_price(value, label):
    if isinstance(value, bool):
        raise ValueError(f'{label}: укажите целое число больше 0.')
    if isinstance(value, int):
        if value <= 0:
            raise ValueError(f'{label} должна быть целым числом больше 0.')
        return value
    if isinstance(value, float):
        if not value.is_integer():
            raise ValueError(f'{label} должна быть целым числом. Дробное значение недопустимо.')
        number = int(value)
        if number <= 0:
            raise ValueError(f'{label} должна быть целым числом больше 0.')
        return number
    text = str(value).strip().replace('\xa0', ' ').replace(' ', '')
    if not re.fullmatch(r'[0-9]+', text):
        raise ValueError(f'{label} должна быть целым числом больше 0.')
    number = int(text)
    if number <= 0:
        raise ValueError(f'{label} должна быть целым числом больше 0.')
    return number


def _parse_stock(value):
    if isinstance(value, bool):
        raise ValueError('Остаток: укажите целое число 0 или больше.')
    if isinstance(value, int):
        if value < 0:
            raise ValueError('Остаток не может быть отрицательным.')
        return value
    if isinstance(value, float):
        if not value.is_integer():
            raise ValueError('Остаток должен быть целым числом.')
        number = int(value)
        if number < 0:
            raise ValueError('Остаток не может быть отрицательным.')
        return number
    text = str(value).strip().replace('\xa0', ' ').replace(' ', '')
    if not re.fullmatch(r'[0-9]+', text):
        raise ValueError('Остаток должен быть целым числом 0 или больше.')
    return int(text)


def _map_headers(header_row):
    mapping = {}
    for index, value in enumerate(header_row):
        key = _header_key(value)
        if key and key not in mapping:
            mapping[key] = index
    return mapping


def _cell_at(row_cells, index):
    if index is None or index >= len(row_cells):
        return None
    return row_cells[index]


def _match_product(seller, article):
    qs = Product.objects.owned_by_seller(seller)
    exact = list(qs.filter(article=article)[:3])
    if len(exact) == 1:
        return exact[0], ''
    if len(exact) > 1:
        return None, 'Несколько товаров с этим артикулом у продавца.'
    insensitive = list(qs.filter(article__iexact=article)[:3])
    if len(insensitive) == 1:
        return insensitive[0], ''
    if len(insensitive) > 1:
        return None, 'Несколько товаров с этим артикулом у продавца.'
    return None, 'Товар с этим артикулом не найден у выбранного продавца.'


def _row_is_blank(row_cells):
    return all(_is_empty_cell(getattr(cell, 'value', None)) for cell in row_cells)


def parse_wholesale_update_workbook(seller, payload: bytes):
    try:
        workbook = load_workbook(BytesIO(payload), data_only=False, read_only=True)
    except Exception as exc:
        raise WholesaleUpdateError('Не удалось прочитать XLSX.') from exc
    sheet = workbook.active
    rows_iter = sheet.iter_rows()
    try:
        header_cells = next(rows_iter)
    except StopIteration as exc:
        workbook.close()
        raise WholesaleUpdateError('В файле нет заголовков.') from exc
    headers = [cell.value for cell in header_cells]
    mapping = _map_headers(headers)
    if HEADER_ARTICLE not in mapping:
        workbook.close()
        raise WholesaleUpdateError('Обязательная колонка «Артикул» не найдена.')

    parsed = []
    data_rows = 0
    for row_number, row_cells in enumerate(rows_iter, start=2):
        if _row_is_blank(row_cells):
            continue
        data_rows += 1
        if data_rows > MAX_DATA_ROWS:
            workbook.close()
            raise WholesaleUpdateError(
                f'В файле больше {MAX_DATA_ROWS} строк. Разбейте файл на части.'
            )
        article_cell = _cell_at(row_cells, mapping[HEADER_ARTICLE])
        retail_cell = _cell_at(row_cells, mapping.get(HEADER_RETAIL))
        wholesale_cell = _cell_at(row_cells, mapping.get(HEADER_WHOLESALE))
        stock_cell = _cell_at(row_cells, mapping.get(HEADER_STOCK))
        item = WholesaleUpdateRow(row_number=row_number, article='', action='')
        changeable = {
            'Артикул': article_cell,
            'Розничная цена': retail_cell,
            'Оптовая цена': wholesale_cell,
            'Остаток': stock_cell,
        }
        for label, cell in changeable.items():
            if cell is not None and _cell_is_formula(cell):
                item.errors.append(
                    f'Колонка «{label}» содержит формулу. Загрузите значения, не формулы.'
                )
        article_value = getattr(article_cell, 'value', None)
        if _is_empty_cell(article_value):
            item.errors.append('Пустой артикул.')
            item.article = ''
        else:
            item.article = str(article_value).strip()
        if not item.errors:
            if retail_cell is not None and not _is_empty_cell(retail_cell.value):
                try:
                    item.retail_new = _parse_positive_price(
                        retail_cell.value, 'Розничная цена'
                    )
                except ValueError as exc:
                    item.errors.append(str(exc))
            if wholesale_cell is not None and not _is_empty_cell(wholesale_cell.value):
                try:
                    item.wholesale_new = _parse_positive_price(
                        wholesale_cell.value, 'Оптовая цена'
                    )
                except ValueError as exc:
                    item.errors.append(str(exc))
            if stock_cell is not None and not _is_empty_cell(stock_cell.value):
                try:
                    item.stock_new = _parse_stock(stock_cell.value)
                except ValueError as exc:
                    item.errors.append(str(exc))
        parsed.append(item)
    workbook.close()
    return parsed


def _json_value(value):
    if value is UNCHANGED:
        return None
    if value is None:
        return None
    return int(value) if isinstance(value, int) else value


def _build_changed_fields(item):
    payload = {'baseline': item.baseline}
    if item.retail_new is not UNCHANGED:
        payload['price'] = {
            'old': item.baseline.get('price'),
            'new': int(item.retail_new),
        }
    if item.wholesale_new is not UNCHANGED:
        payload['wholesale_price'] = {
            'old': item.baseline.get('wholesale_price'),
            'new': int(item.wholesale_new),
        }
    if item.stock_new is not UNCHANGED:
        payload['stock_qty'] = {
            'old': item.baseline.get('stock_qty'),
            'new': int(item.stock_new),
        }
    return payload


def plan_wholesale_update(seller, payload: bytes):
    rows = parse_wholesale_update_workbook(seller, payload)
    article_counts = {}
    for item in rows:
        key = item.article.casefold()
        if key:
            article_counts[key] = article_counts.get(key, 0) + 1

    for item in rows:
        if item.errors:
            item.action = CatalogImportItem.ACTION_ERROR
            continue
        key = item.article.casefold()
        if article_counts.get(key, 0) > 1:
            item.action = CatalogImportItem.ACTION_CONFLICT
            item.errors.append('Артикул повторяется в файле.')
            continue
        product, match_error = _match_product(seller, item.article)
        if product is None:
            item.action = CatalogImportItem.ACTION_CONFLICT
            item.errors.append(match_error)
            continue
        item.product = product
        item.baseline = {
            'price': int(product.price) if product.price is not None else None,
            'wholesale_price': min_qty_one_wholesale_price(product),
            'stock_qty': int(product.stock_qty) if product.stock_qty is not None else None,
        }
        retail_old = item.baseline['price']
        wholesale_old = item.baseline['wholesale_price']
        stock_old = item.baseline['stock_qty']
        retail_effective = (
            item.retail_new if item.retail_new is not UNCHANGED else retail_old
        )
        wholesale_effective = (
            item.wholesale_new if item.wholesale_new is not UNCHANGED else wholesale_old
        )
        if (
            item.wholesale_new is not UNCHANGED
            and retail_effective is not None
            and int(wholesale_effective) >= int(retail_effective)
        ):
            item.warnings.append(WHOLESALE_GTE_RETAIL_WARNING)
        changed = False
        if item.retail_new is not UNCHANGED and item.retail_new != retail_old:
            changed = True
        if item.wholesale_new is not UNCHANGED and item.wholesale_new != wholesale_old:
            changed = True
        if item.stock_new is not UNCHANGED and item.stock_new != stock_old:
            changed = True
        item.changed_fields = _build_changed_fields(item)
        if changed:
            item.action = CatalogImportItem.ACTION_UPDATED
        else:
            item.action = CatalogImportItem.ACTION_UNCHANGED
            if (
                item.retail_new is UNCHANGED
                and item.wholesale_new is UNCHANGED
                and item.stock_new is UNCHANGED
            ):
                item.changed_fields = {'baseline': item.baseline}
    return rows


def summarize_rows(rows):
    summary = {
        'rows': len(rows),
        'matched': 0,
        'updated': 0,
        'unchanged': 0,
        'conflicts': 0,
        'errors': 0,
        'retail_changes': 0,
        'wholesale_changes': 0,
        'stock_changes': 0,
        'has_blockers': False,
        'has_warnings': False,
    }
    for item in rows:
        if item.product is not None:
            summary['matched'] += 1
        if item.action == CatalogImportItem.ACTION_UPDATED:
            summary['updated'] += 1
            fields = item.changed_fields or {}
            if 'price' in fields:
                summary['retail_changes'] += 1
            if 'wholesale_price' in fields:
                summary['wholesale_changes'] += 1
            if 'stock_qty' in fields:
                summary['stock_changes'] += 1
        elif item.action == CatalogImportItem.ACTION_UNCHANGED:
            summary['unchanged'] += 1
        elif item.action == CatalogImportItem.ACTION_CONFLICT:
            summary['conflicts'] += 1
        elif item.action == CatalogImportItem.ACTION_ERROR:
            summary['errors'] += 1
        if item.warnings:
            summary['has_warnings'] = True
    summary['has_blockers'] = bool(summary['conflicts'] or summary['errors'])
    return summary


def persist_wholesale_update_batch(
    *,
    seller,
    rows,
    filename,
    file_sha256,
    mode,
    uploaded_by=None,
    applied_by=None,
    status=None,
):
    summary = summarize_rows(rows)
    now = timezone.now()
    if status is None:
        status = CatalogImportBatch.STATUS_SUCCESS
    batch = CatalogImportBatch.objects.create(
        seller_profile=seller,
        source=CatalogImportBatch.SOURCE_WHOLESALE_UPDATE,
        filename=filename,
        file_sha256=file_sha256,
        started_at=now,
        finished_at=now,
        mode=mode,
        source_scope=CatalogImportBatch.SCOPE_PARTIAL,
        status=status,
        source_row_count=summary['rows'],
        source_unique_count=summary['matched'] + summary['conflicts'] + summary['errors'],
        selected_count=summary['rows'],
        created_count=0,
        updated_count=summary['updated'],
        unchanged_count=summary['unchanged'],
        skipped_count=0,
        conflict_count=summary['conflicts'],
        warning_count=sum(1 for item in rows if item.warnings),
        error_count=summary['errors'],
        missing_from_source_count=0,
        uploaded_by=uploaded_by,
        applied_by=applied_by,
    )
    item_rows = []
    for item in rows:
        changed = dict(item.changed_fields or {})
        if mode == CatalogImportBatch.MODE_WRITE:
            changed.pop('baseline', None)
        item_rows.append(
            CatalogImportItem(
                batch=batch,
                product=item.product,
                article=item.article or '',
                action=item.action,
                warnings=list(item.warnings or []),
                errors=list(item.errors or []),
                changed_fields=changed,
            )
        )
    if item_rows:
        CatalogImportItem.objects.bulk_create(item_rows)
    return batch, summary


def preview_display_rows(batch):
    rows = []
    for item in batch.items.select_related('product').all():
        fields = item.changed_fields or {}
        baseline = fields.get('baseline') or {}
        price = fields.get('price') or {}
        wholesale = fields.get('wholesale_price') or {}
        stock = fields.get('stock_qty') or {}
        rows.append({
            'article': item.article,
            'title': getattr(item.product, 'title', '') or '—',
            'retail_old': price.get('old', baseline.get('price')),
            'retail_new': price.get('new', baseline.get('price')),
            'wholesale_old': wholesale.get('old', baseline.get('wholesale_price')),
            'wholesale_new': wholesale.get('new', baseline.get('wholesale_price')),
            'stock_old': stock.get('old', baseline.get('stock_qty')),
            'stock_new': stock.get('new', baseline.get('stock_qty')),
            'action': item.action,
            'action_label': item.get_action_display(),
            'errors': item.errors or [],
            'warnings': item.warnings or [],
        })
    return rows


def _current_baseline(product):
    return {
        'price': int(product.price) if product.price is not None else None,
        'wholesale_price': min_qty_one_wholesale_price(product),
        'stock_qty': int(product.stock_qty) if product.stock_qty is not None else None,
    }


def _apply_row(item, now):
    product = item.product
    fields = item.changed_fields or {}
    update_fields = []
    if 'price' in fields:
        product.price = int(fields['price']['new'])
        update_fields.append('price')
    if 'stock_qty' in fields:
        product.stock_qty = int(fields['stock_qty']['new'])
        update_fields.append('stock_qty')
    if update_fields:
        product.save(update_fields=update_fields)
    if 'wholesale_price' in fields:
        new_price = int(fields['wholesale_price']['new'])
        tier = ProductPriceTier.objects.filter(product=product, min_qty=1).first()
        if tier is None:
            ProductPriceTier.objects.create(
                product=product,
                min_qty=1,
                price=new_price,
                is_active=True,
            )
        else:
            tier.price = new_price
            tier.is_active = True
            tier.save(update_fields=['price', 'is_active'])
    if 'stock_qty' in fields:
        fulfillment, created = ProductFulfillment.objects.get_or_create(
            product=product,
            defaults={
                'source': ProductFulfillment.SOURCE_MANUAL,
                'external_id': '',
                'last_synced_at': now,
            },
        )
        if not created:
            fulfillment.last_synced_at = now
            fulfillment.save(update_fields=['last_synced_at'])


def apply_wholesale_update_preview(*, seller, preview_batch, applied_by):
    if preview_batch.seller_profile_id != seller.pk:
        raise WholesaleUpdateBlockedError('Пакет принадлежит другому продавцу.')
    if preview_batch.source != CatalogImportBatch.SOURCE_WHOLESALE_UPDATE:
        raise WholesaleUpdateBlockedError('Это не пакет обновления опта.')
    if preview_batch.mode != CatalogImportBatch.MODE_DRY_RUN:
        raise WholesaleUpdateBlockedError('Применить можно только предварительную проверку.')
    items = list(preview_batch.items.select_related('product').all())
    if any(
        item.action in {
            CatalogImportItem.ACTION_CONFLICT,
            CatalogImportItem.ACTION_ERROR,
        }
        for item in items
    ):
        raise WholesaleUpdateBlockedError(
            'В предварительной проверке есть конфликты или ошибки.'
        )

    now = timezone.now()
    write_rows = []
    with transaction.atomic():
        product_ids = sorted(
            {item.product_id for item in items if item.product_id}
        )
        locked = {
            product.pk: product
            for product in Product.objects.select_for_update()
            .filter(pk__in=product_ids)
            .prefetch_related('price_tiers')
        }
        for item in items:
            if item.product_id is None:
                continue
            product = locked.get(item.product_id)
            if product is None:
                raise WholesaleUpdateStaleError(STALE_APPLY_MESSAGE)
            expected = (item.changed_fields or {}).get('baseline') or {}
            current = _current_baseline(product)
            if (
                current.get('price') != expected.get('price')
                or current.get('wholesale_price') != expected.get('wholesale_price')
                or current.get('stock_qty') != expected.get('stock_qty')
            ):
                raise WholesaleUpdateStaleError(STALE_APPLY_MESSAGE)
            if item.action == CatalogImportItem.ACTION_UPDATED:
                item.product = product
                _apply_row(item, now)
            write_item = WholesaleUpdateRow(
                row_number=0,
                article=item.article,
                action=item.action,
                product=product,
                errors=list(item.errors or []),
                warnings=list(item.warnings or []),
                changed_fields=dict(item.changed_fields or {}),
            )
            write_rows.append(write_item)
        write_batch, summary = persist_wholesale_update_batch(
            seller=seller,
            rows=write_rows,
            filename=preview_batch.filename,
            file_sha256=preview_batch.file_sha256,
            mode=CatalogImportBatch.MODE_WRITE,
            uploaded_by=preview_batch.uploaded_by,
            applied_by=applied_by,
            status=CatalogImportBatch.STATUS_SUCCESS,
        )
    return write_batch, summary
