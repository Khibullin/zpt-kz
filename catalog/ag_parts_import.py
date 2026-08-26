"""AG Parts Excel/photo import helpers.

The command inspects real workbooks at runtime. Column aliases below are
fallbacks used after header detection, not a guessed sheet layout.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import zipfile
from dataclasses import dataclass, field, asdict
from pathlib import Path

from django.core.files.base import ContentFile
from django.db import transaction

from catalog.applicability import serialize_plain_list
from catalog.models import (
    Brand,
    CarModel,
    Category,
    Product,
    ProductImage,
    SellerProfile,
)


IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.webp', '.bmp', '.tif', '.tiff'}

ARTICLE_HEADERS = (
    'артикул', 'article', 'oem', 'oe', 'part number', 'part no', 'partno',
    'p/n', 'pn', 'код', 'код товара', 'sku', 'продукт',
)
TITLE_HEADERS = (
    'название', 'наименование', 'name', 'title', 'товар', 'номенклатура',
    'описание товара',
)
CATEGORY_HEADERS = (
    'категория', 'category', 'тип', 'type', 'группа', 'group', 'вид',
)
BRAND_HEADERS = (
    'марка', 'бренд авто', 'бренд', 'brand', 'make', 'марка авто',
)
MODEL_HEADERS = (
    'модель', 'model', 'модели', 'car model', 'модель авто',
)
COMPAT_HEADERS = (
    'применимость', 'совместимость', 'compatibility', 'application',
    'fitment', 'подходит', 'применяемость',
)
DESCRIPTION_HEADERS = (
    'описание', 'description', 'desc', 'текст описания',
)
RETAIL_HEADERS = (
    'розница', 'розничная цена', 'цена розница', 'retail', 'retail price',
    'price', 'цена', 'kaspi', 'каспи',
)
COST_HEADERS = (
    'себестоимость', 'с.ст', 'с.ст.', 'sst', 'cost', 'cost_price',
    'себест', 'закуп', 'закупка цена',
)
QTY_HEADERS = (
    'кол-во', 'количество', 'qty', 'quantity', 'закуплено', 'остаток',
    'stock', 'закупка',
)
ADDITIONAL_MODEL_HEADERS = (
    'дополнительные модели', 'доп. модели', 'доп модели',
    'additional models', 'extra models', 'selected models',
)
ENGINE_HEADERS = (
    'двигатели', 'двигатель', 'engines', 'engine',
    'engine compatibility', 'двигатели применяемость',
)
OEM_CROSS_HEADERS = (
    'oem / кросс-номера', 'oem/кросс-номера', 'oem / кросс номера',
    'кросс-номера', 'кросс номера', 'cross references',
    'cross-references', 'oem cross', 'oem кросс', 'кроссы',
)

FILTER_TYPE_KEYS = {
    'air filter': 'Воздушный фильтр',
    'cabin filter': 'Салонный фильтр',
    'caibin filter': 'Салонный фильтр',
    'oil filter': 'Масляный фильтр',
    'fuel filter': 'Топливный фильтр',
    'воздушный фильтр': 'Воздушный фильтр',
    'салонный фильтр': 'Салонный фильтр',
    'салонные фильтры': 'Салонный фильтр',
    'масляный фильтр': 'Масляный фильтр',
    'масляные фильтры': 'Масляный фильтр',
    'маслянные фильтры': 'Масляный фильтр',
    'масляные фильтры двигателя': 'Масляный фильтр',
    'маслянные фильтры двигателя': 'Масляный фильтр',
    'топливный фильтр': 'Топливный фильтр',
}
SPARK_TYPE_KEYS = {
    'spark plug': 'Свеча зажигания',
    'свеча зажигания': 'Свеча зажигания',
    'свечи зажигания': 'Свеча зажигания',
    'свечи': 'Свеча зажигания',
}
PHOTO_FOLDER_TYPES = {
    'салонные фильтры': ('Фильтры', 'Салонный фильтр'),
    'маслянные фильтры двигателя': ('Фильтры', 'Масляный фильтр'),
    'масляные фильтры двигателя': ('Фильтры', 'Масляный фильтр'),
    'свечи': ('Электрика', 'Свеча зажигания'),
}

BRAND_ALIASES = {
    'chery': 'Chery',
    'haval': 'Haval',
    'geely': 'Geely',
    'changan': 'Changan',
    'byd': 'BYD',
    'exeed': 'Exeed',
    'omoda': 'Omoda',
    'jaecoo': 'Jaecoo',
    'tank': 'Tank',
    'gac': 'GAC',
    'jetour': 'Jetour',
    'kaiyi': 'Kaiyi',
    'lixiang': 'Lixiang',
    'li': 'Lixiang',
    'zeekr': 'Zeekr',
    'toyota': 'Toyota',
    'hyundai': 'Hyundai',
    'kia': 'Kia',
    'chevrolet': 'Chevrolet',
    'nissan': 'Nissan',
    'bmw': 'BMW',
    'mercedes': 'Mercedes-Benz',
    'mercedes-benz': 'Mercedes-Benz',
    'lexus': 'Lexus',
    'jeep': 'Jeep',
    'great wall': 'Great Wall',
    'gw': 'Great Wall',
}



def normalize_header(value):
    text = ' '.join(str(value or '').strip().lower().replace('\n', ' ').split())
    text = text.replace('ё', 'е')
    return text


def cell_text(value):
    if value is None:
        return ''
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).rstrip('0').rstrip('.')
    return str(value).strip()


def parse_money(value):
    text = cell_text(value)
    if not text:
        return None
    cleaned = (
        text.replace('\xa0', ' ')
        .replace(' ', '')
        .replace('₸', '')
        .replace('тг', '')
        .replace('tg', '')
        .replace(',', '.')
    )
    cleaned = re.sub(r'[^0-9.]', '', cleaned)
    if not cleaned:
        return None
    try:
        amount = float(cleaned)
    except ValueError:
        return None
    if amount <= 0:
        return None
    return int(round(amount))


def normalize_article(value):
    text = cell_text(value)
    return re.sub(r'[^A-Za-z0-9]', '', text).upper()


def display_article(value):
    return cell_text(value)


def extract_article(value):
    text = cell_text(value)
    if not text:
        return '', ''
    if re.fullmatch(r'[A-Za-z0-9][A-Za-z0-9._/-]*', text):
        return text, normalize_article(text)
    tokens = [token for token in re.split(r'\s+', text) if token]
    candidate = tokens[-1] if tokens else text
    return candidate, normalize_article(candidate)


def resolve_product_type(raw):
    key = normalize_header(raw)
    if not key:
        return '', ''
    if key in FILTER_TYPE_KEYS:
        return 'Фильтры', FILTER_TYPE_KEYS[key]
    if key in SPARK_TYPE_KEYS:
        return 'Электрика', SPARK_TYPE_KEYS[key]
    for prefix, label in FILTER_TYPE_KEYS.items():
        if prefix in key:
            return 'Фильтры', label
    for prefix, label in SPARK_TYPE_KEYS.items():
        if prefix in key:
            return 'Электрика', label
    return '', ''


def type_from_photo_path(path):
    for part in Path(path).parts:
        mapped = PHOTO_FOLDER_TYPES.get(normalize_header(part))
        if mapped:
            return mapped
    return '', ''


def alias_brand_name(raw):
    key = normalize_header(raw)
    return BRAND_ALIASES.get(key, (raw or '').strip())


def detect_column_map(headers):
    mapping = {}
    used = set()
    groups = {
        'engine_compatibility': ENGINE_HEADERS,
        'oem_cross_references': OEM_CROSS_HEADERS,
        'article': ARTICLE_HEADERS,
        'title': TITLE_HEADERS,
        'category': CATEGORY_HEADERS,
        'brand': BRAND_HEADERS,
        'model': MODEL_HEADERS,
        'compatibility': COMPAT_HEADERS,
        'description': DESCRIPTION_HEADERS,
        'retail_price': RETAIL_HEADERS,
        'cost_price': COST_HEADERS,
        'quantity': QTY_HEADERS,
        'extra_models': ADDITIONAL_MODEL_HEADERS,
    }
    normalized = [normalize_header(h) for h in headers]
    for field, aliases in groups.items():
        for index, header in enumerate(normalized):
            if index in used or not header:
                continue
            if header in aliases or any(
                header == alias or header.startswith(alias + ' ')
                for alias in aliases
            ):
                mapping[field] = index
                used.add(index)
                break
    return mapping


@dataclass
class SheetInspection:
    name: str
    headers: list
    column_map: dict
    sample_rows: list
    row_count: int
    embedded_image_count: int
    embedded_image_rows: list
    quantity_columns: list


@dataclass
class WorkbookInspection:
    path: str
    sheets: list
    chosen_sheet: str
    notes: list = field(default_factory=list)


@dataclass
class PhotoMatch:
    path: str
    article_key: str


@dataclass
class PreparedRow:
    article: str
    article_key: str
    title: str
    category_raw: str
    category_name: str
    brand_raw: str
    model_raw: str
    compatibility: str
    retail_price: int | None
    cost_price: int | None
    quantity_raw: str
    product_type: str = ''
    description: str = ''
    extra_models_raw: str = ''
    engine_compatibility: str = ''
    oem_cross_references: str = ''
    photos: list = field(default_factory=list)
    source_row: int = 0
    source_sheet: str = ''
    warnings: list = field(default_factory=list)
    errors: list = field(default_factory=list)


@dataclass
class ImportResult:
    article: str
    action: str
    source_row: int
    warnings: list = field(default_factory=list)
    errors: list = field(default_factory=list)
    product_id: int | None = None


def _iter_data_rows(worksheet):
    rows = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return [], []
    headers = [cell_text(value) for value in header_row]
    data = []
    for index, row in enumerate(rows, start=2):
        values = list(row)
        if not any(cell_text(value) for value in values):
            continue
        data.append((index, values))
    return headers, data


def _embedded_image_rows(worksheet):
    rows = []
    for image in getattr(worksheet, '_images', []) or []:
        anchor = getattr(image, 'anchor', None)
        source = getattr(anchor, '_from', None)
        if source is None:
            continue
        rows.append(int(source.row) + 1)
    return rows


def inspect_workbook(path):
    from openpyxl import load_workbook

    path = Path(path)
    workbook = load_workbook(path, data_only=True)
    sheets = []
    notes = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        headers, data_rows = _iter_data_rows(worksheet)
        column_map = detect_column_map(headers)
        qty_cols = []
        for index, header in enumerate(headers):
            if normalize_header(header) in QTY_HEADERS:
                qty_cols.append(header)
        sample = []
        for row_number, values in data_rows[:5]:
            sample.append({
                'row': row_number,
                'values': [cell_text(value) for value in values[:20]],
            })
        image_rows = _embedded_image_rows(worksheet)
        sheets.append(SheetInspection(
            name=sheet_name,
            headers=headers,
            column_map=column_map,
            sample_rows=sample,
            row_count=len(data_rows),
            embedded_image_count=len(getattr(worksheet, '_images', []) or []),
            embedded_image_rows=sorted(set(image_rows)),
            quantity_columns=qty_cols,
        ))
    workbook.close()

    chosen = ''
    if sheets:
        with_article = [item for item in sheets if 'article' in item.column_map]
        chosen = (with_article[0].name if with_article else sheets[0].name)
    if not chosen:
        notes.append('workbook_has_no_sheets')
    return WorkbookInspection(
        path=str(path),
        sheets=sheets,
        chosen_sheet=chosen,
        notes=notes,
    )


def load_sheet_rows(path, sheet_name=None):
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    headers, data_rows = _iter_data_rows(worksheet)
    column_map = detect_column_map(headers)
    images_by_row = {}
    for image in getattr(worksheet, '_images', []) or []:
        anchor = getattr(image, 'anchor', None)
        source = getattr(anchor, '_from', None)
        if source is None:
            continue
        row_number = int(source.row) + 1
        payload = None
        if hasattr(image, '_data'):
            try:
                payload = image._data()
            except Exception:
                payload = None
        if not payload:
            continue
        images_by_row.setdefault(row_number, []).append(payload)
    workbook.close()
    return headers, column_map, data_rows, images_by_row


def _row_value(values, column_map, field):
    index = column_map.get(field)
    if index is None or index >= len(values):
        return ''
    return values[index]


def merge_prepared(existing, incoming):
    if incoming.title and not existing.title:
        existing.title = incoming.title
    if incoming.category_raw and not existing.category_raw:
        existing.category_raw = incoming.category_raw
        existing.category_name = incoming.category_name
    if incoming.brand_raw and not existing.brand_raw:
        existing.brand_raw = incoming.brand_raw
    if incoming.model_raw and not existing.model_raw:
        existing.model_raw = incoming.model_raw
    if incoming.compatibility:
        if existing.compatibility and incoming.compatibility not in existing.compatibility:
            existing.compatibility = f'{existing.compatibility}; {incoming.compatibility}'
        elif not existing.compatibility:
            existing.compatibility = incoming.compatibility
    if existing.retail_price is None:
        existing.retail_price = incoming.retail_price
    if existing.cost_price is None:
        existing.cost_price = incoming.cost_price
    if incoming.product_type and not existing.product_type:
        existing.product_type = incoming.product_type
        existing.category_name = incoming.category_name
    if incoming.description and not existing.description:
        existing.description = incoming.description
    if incoming.extra_models_raw and not existing.extra_models_raw:
        existing.extra_models_raw = incoming.extra_models_raw
    if incoming.engine_compatibility and not existing.engine_compatibility:
        existing.engine_compatibility = incoming.engine_compatibility
    if incoming.oem_cross_references and not existing.oem_cross_references:
        existing.oem_cross_references = incoming.oem_cross_references
    existing.photos.extend(incoming.photos)
    existing.warnings.extend(incoming.warnings)
    existing.warnings.append(
        f'duplicate_source_row:{incoming.source_sheet}:{incoming.source_row}'
    )
    return existing


def prepared_from_excel_row(row_number, values, column_map, sheet_name, images_by_row):
    raw_article = _row_value(values, column_map, 'article')
    article, article_key = extract_article(raw_article)
    warnings = []
    errors = []
    if not article_key:
        errors.append('missing_article')
    if cell_text(raw_article) and article != cell_text(raw_article):
        warnings.append(f'article_extracted_from:{cell_text(raw_article)}')
    category_raw = cell_text(_row_value(values, column_map, 'category'))
    zpt_category, product_type = resolve_product_type(category_raw or cell_text(raw_article))
    retail_raw = _row_value(values, column_map, 'retail_price')
    retail_price = parse_money(retail_raw)
    if 'retail_price' not in column_map:
        warnings.append('missing_retail_price')
    elif not cell_text(retail_raw):
        warnings.append('missing_retail_price')
    elif not retail_price:
        warnings.append('invalid_retail_price')
    photos = []
    for payload in images_by_row.get(row_number, []):
        photos.append({
            'kind': 'embedded',
            'payload': payload,
            'name': f'{article_key or "row"}-{row_number}.png',
        })
    if len(images_by_row.get(row_number, [])) > 3:
        warnings.append('many_embedded_images_on_row')
    return PreparedRow(
        article=article,
        article_key=article_key,
        title=cell_text(_row_value(values, column_map, 'title')),
        category_raw=category_raw,
        category_name=zpt_category,
        product_type=product_type,
        brand_raw=cell_text(_row_value(values, column_map, 'brand')),
        model_raw=cell_text(_row_value(values, column_map, 'model')),
        compatibility=cell_text(_row_value(values, column_map, 'compatibility')),
        extra_models_raw=cell_text(_row_value(values, column_map, 'extra_models')),
        engine_compatibility=serialize_plain_list(
            _row_value(values, column_map, 'engine_compatibility')
        ),
        oem_cross_references=serialize_plain_list(
            _row_value(values, column_map, 'oem_cross_references')
        ),
        description=cell_text(_row_value(values, column_map, 'description')),
        retail_price=retail_price,
        cost_price=parse_money(_row_value(values, column_map, 'cost_price')),
        quantity_raw=cell_text(_row_value(values, column_map, 'quantity')),
        photos=photos,
        source_row=row_number,
        source_sheet=sheet_name,
        warnings=warnings,
        errors=errors,
    )


def load_cost_index(path):
    inspection = inspect_workbook(path)
    headers, column_map, data_rows, _images = load_sheet_rows(
        path,
        inspection.chosen_sheet,
    )
    index = {}
    warnings = []
    if 'article' not in column_map:
        warnings.append('cost_workbook_missing_article_column')
        return index, inspection, warnings
    for row_number, values in data_rows:
        article_key = normalize_article(_row_value(values, column_map, 'article'))
        cost = parse_money(_row_value(values, column_map, 'cost_price'))
        if not article_key:
            continue
        if article_key in index and index[article_key] != cost:
            warnings.append(f'cost_conflict:{article_key}')
        index[article_key] = cost
    return index, inspection, warnings


def unpack_archives(archives, dest_dir):
    dest = Path(dest_dir)
    dest.mkdir(parents=True, exist_ok=True)
    unpacked = []
    for archive in archives:
        archive_path = Path(archive)
        if archive_path.suffix.lower() != '.zip':
            raise ValueError(f'unsupported_archive:{archive_path}')
        with zipfile.ZipFile(archive_path) as bundle:
            bundle.extractall(dest / archive_path.stem)
        unpacked.append(dest / archive_path.stem)
    return unpacked


def photo_article_keys(path, search_root):
    rel = path.relative_to(search_root)
    tokens = [path.stem, *[part for part in rel.parts[:-1]]]
    keys = set()
    for token in tokens:
        keys.add(normalize_article(token))
        parts = re.split(r'[-_\s]+', str(token))
        acc = ''
        for part in parts:
            if not part:
                continue
            acc += part
            keys.add(normalize_article(acc))
    return {key for key in keys if key}


def index_photos(photo_roots):
    by_article = {}
    total = 0
    for root in photo_roots:
        root_path = Path(root)
        if not root_path.exists():
            continue
        if root_path.is_file() and root_path.suffix.lower() in IMAGE_EXTENSIONS:
            files = [root_path]
            search_root = root_path.parent
        else:
            files = [
                path for path in root_path.rglob('*')
                if path.is_file() and path.suffix.lower() in IMAGE_EXTENSIONS
            ]
            search_root = root_path
        for path in sorted(files):
            keys = photo_article_keys(path, search_root)
            if not keys:
                continue
            total += 1
            for key in keys:
                by_article.setdefault(key, []).append(str(path))
    for key, paths in by_article.items():
        unique = []
        seen = set()
        for item in paths:
            digest = hashlib.md5(Path(item).read_bytes()).hexdigest()
            if digest in seen:
                continue
            seen.add(digest)
            unique.append(item)
        by_article[key] = unique
    return by_article, total


def attach_archive_photos(rows, photo_index):
    for row in rows:
        if not row.article_key:
            continue
        matches = photo_index.get(row.article_key, [])
        for path in matches:
            row.photos.append({
                'kind': 'file',
                'path': path,
                'name': Path(path).name,
            })
        seen = set()
        unique = []
        for photo in row.photos:
            marker = photo.get('path') or photo.get('name')
            if marker in seen:
                continue
            seen.add(marker)
            unique.append(photo)
        row.photos = unique
        if not row.photos:
            row.warnings.append('NO_LOCAL_PHOTO')
        else:
            folder_category, folder_type = type_from_photo_path(row.photos[0]['path'])
            if folder_category and not row.category_name:
                row.category_name = folder_category
            if folder_type and not row.product_type:
                row.product_type = folder_type


def parse_brand_model_pairs(raw):
    """Parse optional 'Brand:Model; Brand:Model' additional applicability."""
    pairs = []
    warnings = []
    for chunk in re.split(r'[;\n]+', raw or ''):
        chunk = chunk.strip()
        if not chunk:
            continue
        if ':' not in chunk:
            warnings.append(f'invalid_additional_model:{chunk}')
            continue
        brand_raw, model_raw = chunk.split(':', 1)
        brand_raw = brand_raw.strip()
        model_raw = model_raw.strip()
        if not brand_raw or not model_raw:
            warnings.append(f'invalid_additional_model:{chunk}')
            continue
        pairs.append((brand_raw, model_raw))
    return pairs, warnings


class CatalogMatcher:
    def __init__(self):
        self.categories = {
            normalize_header(item.name): item
            for item in Category.objects.all()
        }
        self.brands = {}
        for brand in Brand.objects.select_related('country'):
            self.brands.setdefault(normalize_header(brand.name), []).append(brand)
        self.models = {}
        for model in CarModel.objects.select_related('brand'):
            key = (model.brand_id, normalize_header(model.name))
            self.models.setdefault(key, model)
        self.models_by_brand = {}
        for model in CarModel.objects.select_related('brand'):
            self.models_by_brand.setdefault(model.brand_id, []).append(model)

    def category(self, mapped_name):
        if not mapped_name:
            return None, 'missing_category'
        found = self.categories.get(normalize_header(mapped_name))
        if found:
            return found, None
        return None, f'unknown_category:{mapped_name}'

    def brand(self, raw_name):
        aliased = alias_brand_name(raw_name)
        if not aliased:
            return None, None
        matches = self.brands.get(normalize_header(aliased), [])
        if len(matches) == 1:
            return matches[0], None
        if len(matches) > 1:
            return None, f'ambiguous_brand:{aliased}'
        return None, f'unknown_brand:{raw_name}'

    def model(self, brand, raw_name):
        if not brand or not raw_name:
            return None, None
        found = self.models.get((brand.id, normalize_header(raw_name)))
        if found:
            return found, None
        return None, f'unknown_model:{brand.name}:{raw_name}'

    def parse_fitment(self, row):
        warnings = []
        brands = []
        models = []

        def add_pair(brand_raw, model_raw):
            brand, warning = self.brand(brand_raw)
            if warning:
                warnings.append(warning)
            if brand and brand not in brands:
                brands.append(brand)
            if not model_raw:
                return brand, None
            if not brand:
                return None, None
            model, model_warning = self.model(brand, model_raw)
            if model_warning:
                warnings.append(model_warning)
            if model and model not in models:
                models.append(model)
            return brand, model

        primary_brand = None
        primary_model = None
        if row.brand_raw:
            tokens = [
                token.strip()
                for token in re.split(r'[,/;]+', row.model_raw or '')
                if token.strip()
            ]
            primary_brand, primary_model = add_pair(
                row.brand_raw,
                tokens[0] if tokens else '',
            )
            for token in tokens[1:]:
                add_pair(row.brand_raw, token)

        extra_pairs, extra_warnings = parse_brand_model_pairs(row.extra_models_raw)
        warnings.extend(extra_warnings)
        for brand_raw, model_raw in extra_pairs:
            add_pair(brand_raw, model_raw)

        if not primary_brand and (row.brand_raw or extra_pairs):
            if not any(item.startswith('unknown_brand:') for item in warnings):
                warnings.append('unmatched_fitment')
        return primary_brand, primary_model, brands, models, warnings


def build_title(row, brand, model, category):
    raw_title = (row.title or '').strip()
    extracted_from_title = raw_title and extract_article(raw_title)[1] == row.article_key
    if raw_title and not extracted_from_title:
        return raw_title[:255]
    type_label = row.product_type or (category.name if category else '')
    bits = [part for part in (type_label, brand.name if brand else '', model.name if model else '') if part]
    core = ' '.join(bits) if bits else (row.article or 'AG Parts')
    if row.article and row.article not in core:
        core = f'{core} — {row.article}'
    return core[:255]


def _stored_stem(name):
    stem = Path(name).stem
    return re.sub(r'_[A-Za-z0-9]{7}$', '', stem)


def _product_has_photo_name(product, filename):
    wanted_stem = Path(filename).stem
    names = []
    if product.main_image:
        names.append(Path(product.main_image.name).name)
    names.extend(Path(item.image.name).name for item in product.images.all())
    for name in names:
        if _stored_stem(name) == wanted_stem or Path(name).stem == wanted_stem:
            return True
    return False


def _photo_content(photo):
    if photo.get('kind') == 'file':
        data = Path(photo['path']).read_bytes()
        name = photo.get('name') or Path(photo['path']).name
        return name, data
    name = photo.get('name') or 'image.png'
    return name, photo['payload']


def apply_images(product, photos, replace_images):
    if not photos:
        return
    if replace_images:
        if product.main_image:
            product.main_image.delete(save=False)
        product.images.all().delete()
        product.main_image = None
        product.save(update_fields=['main_image'])
    existing = bool(product.main_image) or product.images.exists()
    if existing and not replace_images:
        for photo in photos:
            name, _data = _photo_content(photo)
            if _product_has_photo_name(product, name):
                continue
            # Do not add extra gallery files on a normal re-run.
            return
        return
    stored = 0
    for photo in photos:
        name, data = _photo_content(photo)
        if Path(name).suffix.lower() not in IMAGE_EXTENSIONS and photo.get('kind') != 'embedded':
            continue
        if stored == 0 and not product.main_image:
            product.main_image.save(name, ContentFile(data), save=True)
        else:
            if not replace_images and _product_has_photo_name(product, name):
                continue
            gallery = ProductImage(product=product, sort_order=stored)
            gallery.image.save(name, ContentFile(data), save=True)
        stored += 1


def _seller_names_match(left, right):
    return (left or '').strip().casefold() == (right or '').strip().casefold()


def find_legacy_ag_parts(article, seller):
    if seller is None or not article:
        return []
    candidates = []
    for product in Product.objects.filter(
        seller_profile__isnull=True,
        article=article,
    ):
        if _seller_names_match(product.seller_name, seller.name):
            candidates.append(product)
    return candidates


def other_seller_products(article, seller, *, exclude_ids=None):
    exclude_ids = set(exclude_ids or [])
    queryset = Product.objects.filter(article=article)
    if seller is not None:
        queryset = queryset.exclude(seller_profile=seller)
    others = []
    for product in queryset:
        if product.pk in exclude_ids:
            continue
        if (
            seller is not None
            and product.seller_profile_id is None
            and _seller_names_match(product.seller_name, seller.name)
        ):
            continue
        others.append(product)
    return others


def _apply_row_fields(product, *, title, category, brand, model, row, seller, description, retail_ok):
    product.title = title
    product.category = category
    product.brand = brand
    product.car_model = model
    product.compatibility = row.compatibility or product.compatibility
    if row.engine_compatibility:
        product.engine_compatibility = row.engine_compatibility
    if row.oem_cross_references:
        product.oem_cross_references = row.oem_cross_references
    if description:
        product.description = description
    if row.cost_price is not None:
        product.cost_price = row.cost_price
    if retail_ok:
        product.price = row.retail_price
        product.price_on_request = False
    if seller.name:
        product.seller_name = seller.name
    if seller.phone:
        product.whatsapp_number = seller.phone
    if seller.city:
        product.city = seller.city
    return product


def upsert_product(
    row,
    seller,
    matcher,
    *,
    dry_run,
    replace_images,
    expect_cost=False,
    adopt_legacy=False,
):
    result = ImportResult(
        article=row.article,
        action='skipped',
        source_row=row.source_row,
        warnings=list(row.warnings),
        errors=list(row.errors),
    )
    if row.errors:
        result.action = 'error'
        return result

    existing = None
    legacy_candidates = []
    if seller is not None:
        phaeton = Product.objects.filter(
            article=row.article,
            supplier=Product.SUPPLIER_PHAETON,
        )
        if phaeton.exists() and not Product.objects.filter(
            seller_profile=seller,
            article=row.article,
        ).exists():
            result.warnings.append('phaeton_article_exists_left_untouched')

        existing = Product.objects.filter(
            seller_profile=seller,
            article=row.article,
        ).first()
        if existing is None:
            legacy_candidates = find_legacy_ag_parts(row.article, seller)
        others = other_seller_products(
            row.article,
            seller,
            exclude_ids=[item.pk for item in legacy_candidates],
        )
        if others:
            result.warnings.append('same_article_exists_for_other_seller')

    if existing and existing.supplier == Product.SUPPLIER_PHAETON:
        result.action = 'skipped'
        result.errors.append('refusing_to_modify_phaeton_product')
        result.product_id = existing.pk
        return result

    if existing is None and len(legacy_candidates) > 1:
        result.action = 'legacy_ag_parts_ambiguous'
        result.errors.append('legacy_ag_parts_ambiguous')
        result.warnings.append('LEGACY_AG_PARTS_AMBIGUOUS')
        return result

    legacy = legacy_candidates[0] if len(legacy_candidates) == 1 else None

    category, category_warning = matcher.category(row.category_name)
    if category_warning:
        result.warnings.append(category_warning)

    brand, model, brands, models, fitment_warnings = matcher.parse_fitment(row)
    result.warnings.extend(fitment_warnings)
    if expect_cost and not row.cost_price:
        result.warnings.append('missing_cost_price')
    if row.quantity_raw:
        result.warnings.append(f'quantity_ignored:{row.quantity_raw}')

    retail_ok = bool(row.retail_price and row.retail_price > 0)
    title = build_title(row, brand, model, category)
    description = (row.description or '').strip()

    if existing is None and legacy is not None:
        result.product_id = legacy.pk
        result.warnings.append(f'LEGACY_AG_PARTS_MATCH product_id={legacy.pk}')
        if dry_run:
            result.action = 'would_adopt' if adopt_legacy else 'legacy_ag_parts_match'
            return result
        if not adopt_legacy:
            result.action = 'skipped'
            result.warnings.append('legacy_requires_adoption')
            return result
        if seller is None:
            result.action = 'error'
            result.errors.append('seller_profile_required_for_write')
            return result
        with transaction.atomic():
            legacy.seller_profile = seller
            _apply_row_fields(
                legacy,
                title=title,
                category=category,
                brand=brand,
                model=model,
                row=row,
                seller=seller,
                description=description,
                retail_ok=retail_ok,
            )
            legacy.save()
            legacy.selected_brands.set(brands)
            legacy.selected_models.set(models)
            apply_images(legacy, row.photos, replace_images=replace_images)
            result.action = 'adopted'
            result.product_id = legacy.pk
        return result

    if dry_run:
        result.action = 'updated' if existing else 'created'
        result.product_id = existing.pk if existing else None
        return result

    if seller is None:
        result.action = 'error'
        result.errors.append('seller_profile_required_for_write')
        return result

    with transaction.atomic():
        if existing:
            _apply_row_fields(
                existing,
                title=title,
                category=category,
                brand=brand,
                model=model,
                row=row,
                seller=seller,
                description=description,
                retail_ok=retail_ok,
            )
            existing.save()
            product = existing
            result.action = 'updated'
        else:
            product = Product.objects.create(
                title=title,
                article=row.article,
                price=row.retail_price if retail_ok else None,
                price_on_request=not retail_ok,
                condition='new',
                status='hidden',
                brand=brand,
                car_model=model,
                category=category,
                seller_name=seller.name,
                whatsapp_number=seller.phone,
                city=seller.city or '',
                seller_profile=seller,
                cost_price=row.cost_price,
                compatibility=row.compatibility,
                engine_compatibility=row.engine_compatibility,
                oem_cross_references=row.oem_cross_references,
                description=description,
                supplier=Product.SUPPLIER_LOCAL,
            )
            result.action = 'created'
        product.selected_brands.set(brands)
        product.selected_models.set(models)
        apply_images(product, row.photos, replace_images=replace_images)
        result.product_id = product.pk
    return result


def filter_rows(rows, *, limit=None, articles=None):
    selected = rows
    if articles:
        wanted = {normalize_article(item) for item in articles if item.strip()}
        selected = [row for row in selected if row.article_key in wanted]
    if limit is not None:
        selected = selected[:limit]
    return selected


def write_reports(results, report_path):
    if not report_path:
        return
    path = Path(report_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = [asdict(item) for item in results]
    path.with_suffix('.json').write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding='utf-8',
    )
    with path.with_suffix('.csv').open('w', encoding='utf-8', newline='') as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=['article', 'action', 'source_row', 'product_id', 'warnings', 'errors'],
        )
        writer.writeheader()
        for item in payload:
            writer.writerow({
                'article': item['article'],
                'action': item['action'],
                'source_row': item['source_row'],
                'product_id': item['product_id'] or '',
                'warnings': '|'.join(item['warnings']),
                'errors': '|'.join(item['errors']),
            })


def summarize(results):
    totals = {
        'CREATED': 0,
        'UPDATED': 0,
        'LEGACY_MATCH': 0,
        'WOULD_ADOPT': 0,
        'ADOPTED': 0,
        'SKIPPED': 0,
        'WARNING': 0,
        'ERROR': 0,
    }
    for item in results:
        if item.action == 'created':
            totals['CREATED'] += 1
        elif item.action == 'updated':
            totals['UPDATED'] += 1
        elif item.action == 'legacy_ag_parts_match':
            totals['LEGACY_MATCH'] += 1
        elif item.action == 'would_adopt':
            totals['LEGACY_MATCH'] += 1
            totals['WOULD_ADOPT'] += 1
        elif item.action == 'adopted':
            totals['ADOPTED'] += 1
        elif item.action == 'skipped':
            totals['SKIPPED'] += 1
            if any('LEGACY_AG_PARTS_MATCH' in warning for warning in item.warnings):
                totals['LEGACY_MATCH'] += 1
        elif item.action in {'error', 'legacy_ag_parts_ambiguous'}:
            totals['ERROR'] += 1
        if item.warnings:
            totals['WARNING'] += 1
        if item.errors and item.action not in {'error', 'legacy_ag_parts_ambiguous'}:
            totals['ERROR'] += 1
    return totals
